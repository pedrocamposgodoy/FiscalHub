# ================================================================
# fiscalhub_app.py — FiscalHub · Portal Asesoría Fiscal
# Nolasco Capital ecosystem
# Diseño: degradados azul/gris pastel · sidebar azul marino
# ================================================================

import streamlit as st
import requests
import pandas as pd
from datetime import datetime, date
import io
import html as _html
from nolasco_styles import inject_global_css
from kpi_renderer import render_kpi_row, render_kpi_grid, ACCENT_F, RED, AMBER, GREEN, GREY

# Paleta determinista por cliente — 8 colores profesionales
# Siempre el mismo color para el mismo cliente_id
_PALETA_CLI = [
    "#1E3A5F",  # azul marino
    "#3D2B6B",  # morado oscuro
    "#1A4731",  # verde oscuro
    "#7A2D1A",  # rojo ladrillo
    "#1A3A4A",  # azul petróleo
    "#4A3000",  # marrón dorado
    "#2D1A4A",  # índigo oscuro
    "#3B3B3B",  # gris antracita
]

def _color_cli(cliente_id: str) -> str:
    """Color determinista para un cliente — siempre el mismo."""
    return _PALETA_CLI[abs(hash(str(cliente_id))) % len(_PALETA_CLI)]

def _e(s):
    """Escapar caracteres HTML especiales en datos de usuario."""
    return _html.escape(str(s)) if s else ""

st.set_page_config(
    page_title="FiscalHub · Nolasco Capital",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded",
)

SUPABASE_URL = "https://odxixtgqcyddfqaapqgi.supabase.co"
SUPABASE_KEY = "sb_publishable_Obgti7yMfXw8wCUL2FbTtA_EWeyHuM9"

def _h(token=None):
    t = token or st.session_state.get("fh_token") or SUPABASE_KEY
    return {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {t}",
            "Content-Type": "application/json", "Prefer": "return=representation"}

def _hd():
    return {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json"}

# ── INYECTAR ESTILO GLOBAL ──────────────────────────────────────
# El CSS completo se gestiona desde nolasco_styles.py (único fuente de verdad)
# Esto se llama UNA VEZ al inicio de la app


# ── Helpers ──────────────────────────────────────────────────────
def sf(v, d=0):
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)): return float(d)
        return float(v)
    except: return float(d)

def _gv(row, *keys, d=0):
    for k in keys:
        v = row.get(k)
        if v is not None:
            try:
                f = float(v)
                if not pd.isna(f): return f
            except: pass
    return float(d)

def fmt_eur(n, sign=False):
    n = float(n or 0)
    s = f"{abs(n):,.0f}".replace(",",".")
    prefix = "−" if n < 0 else ("+" if sign else "")
    return f"{prefix}{s} €"

def days_to_irpf():
    hoy = date.today()
    cierre = date(hoy.year, 6, 30)
    if hoy > cierre: cierre = date(hoy.year+1, 6, 30)
    return (cierre - hoy).days

# ── Auth ─────────────────────────────────────────────────────────
def login_asesor(email, password):
    r = requests.post(f"{SUPABASE_URL}/auth/v1/token?grant_type=password",
        headers={"apikey": SUPABASE_KEY, "Content-Type": "application/json"},
        json={"email": email, "password": password})
    if r.status_code == 200:
        d = r.json()
        return {"ok": True, "user_id": d["user"]["id"],
                "email": d["user"]["email"], "token": d["access_token"]}
    return {"ok": False, "error": r.json().get("error_description", "Error de acceso")}

def registrar_asesor(email, password, nombre, despacho):
    r = requests.post(f"{SUPABASE_URL}/auth/v1/signup",
        headers={"apikey": SUPABASE_KEY, "Content-Type": "application/json"},
        json={"email": email, "password": password})
    if r.status_code == 200:
        uid = r.json().get("id") or r.json().get("user", {}).get("id")
        if uid:
            requests.post(f"{SUPABASE_URL}/rest/v1/asesores", headers=_h(),
                json={"user_id": uid, "nombre": nombre, "despacho": despacho, "email": email})
        return {"ok": True}
    return {"ok": False, "error": r.json().get("error_description", "Error de registro")}

def get_asesor_info(user_id):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/asesores?user_id=eq.{user_id}&select=*", headers=_h())
    if r.status_code == 200 and r.json(): return r.json()[0]
    return {"nombre": "Asesor", "despacho": "Despacho Fiscal", "email": "", "logo_url": None}

def guardar_logo(user_id, logo_bytes, extension="png"):
    """Guarda logo como base64 en tabla user_profiles."""
    import base64 as _b64
    try:
        b64 = _b64.b64encode(logo_bytes).decode("utf-8")
        data = {"user_id": user_id,
                "logo_b64": f"data:image/{extension};base64,{b64}"}
        hdrs = {**_hd(), "Prefer": "resolution=merge-duplicates,return=minimal"}
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/user_profiles?on_conflict=user_id",
            headers=hdrs, json=data, timeout=20)
        return r.status_code in [200, 201, 204]
    except Exception:
        return False

def leer_logo(user_id):
    """Lee logo base64 desde user_profiles y devuelve bytes."""
    import base64 as _b64
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/user_profiles"
            f"?user_id=eq.{user_id}&select=logo_b64",
            headers=_hd(), timeout=10)
        if r.ok and r.json() and r.json()[0].get("logo_b64"):
            b64_str = r.json()[0]["logo_b64"]
            if "," in b64_str:
                b64_str = b64_str.split(",", 1)[1]
            return _b64.b64decode(b64_str)
    except Exception:
        pass
    return None

# ── Supabase data ─────────────────────────────────────────────────
def get_clientes_vinculados(asesor_user_id):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/accesos_asesor"
        f"?asesor_user_id=eq.{asesor_user_id}&activo=eq.true&select=*", headers=_h())
    return r.json() if r.status_code == 200 else []

def get_inmuebles_propietario(pid):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/inmuebles?user_id=eq.{pid}&select=*", headers=_hd())
    return pd.DataFrame(r.json()) if r.status_code == 200 and r.json() else pd.DataFrame()

def get_movimientos_propietario(pid):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/movimientos?user_id=eq.{pid}&select=*", headers=_hd())
    return pd.DataFrame(r.json()) if r.status_code == 200 and r.json() else pd.DataFrame()

def get_perfil_propietario(pid):
    """Lee tipo_cuenta y datos fiscales del propietario desde user_profiles."""
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/user_profiles?user_id=eq.{pid}"
            f"&select=tipo_cuenta,nombre_sociedad,cif_sociedad,nombre_fiscal,nif",
            headers=_hd()
        )
        if r.status_code == 200 and r.json():
            return r.json()[0]
    except Exception:
        pass
    return {"tipo_cuenta": "particular"}


def vincular_propietario(asesor_user_id, codigo):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/accesos_asesor"
        f"?codigo=eq.{codigo.upper()}&activo=eq.true&select=*", headers=_h())
    if r.status_code == 200 and r.json():
        acc = r.json()[0]
        pid = acc.get("propietario_id")
        nombre = acc.get("nombre") or acc.get("nombre_propietario", "Propietario")
        if not pid: return {"ok": False, "error": "Código sin propietario"}
        requests.patch(f"{SUPABASE_URL}/rest/v1/accesos_asesor?id=eq.{acc['id']}",
            headers={**_h(), "Prefer": "return=minimal"}, json={"asesor_user_id": asesor_user_id})
        return {"ok": True, "propietario_id": pid, "nombre": nombre}
    return {"ok": False, "error": "Código no válido o ya utilizado"}

# ── Análisis fiscal ───────────────────────────────────────────────
def calcular_semaforo_inmueble(row):
    """Evalúa un inmueble — devuelve problemas detectados y estado semáforo."""
    problemas = []
    renta     = _gv(row,"renta","Renta")
    ibi       = _gv(row,"ibi_anual","IBI_Anual")
    amort     = _gv(row,"amortizacion_fiscal","Amortizacion_Fiscal")
    seguro    = _gv(row,"seguro_anual","Seguro_Anual")
    comunidad = _gv(row,"comunidad","Comunidad")
    catastral = _gv(row,"valor_catastral","Valor_Catastral")
    precio    = _gv(row,"precio_compra","Precio_Compra")
    tipo      = str(row.get("tipo_arrendamiento") or row.get("Tipo_Arrendamiento","")).lower()
    fecha_str = str(row.get("fecha_inicio_contrato") or row.get("Fecha_Inicio_Contrato","") or "")
    ingresos  = renta * 12

    if amort == 0 and renta > 0:
        if catastral == 0 and precio == 0:
            problemas.append({"tipo":"crit","titulo":"Amortización sin calcular",
                "desc":"Falta valor catastral y precio de compra.","accion":"Solicitar escritura o consultar catastro"})
        else:
            problemas.append({"tipo":"crit","titulo":"Amortización a 0 — revisar",
                "desc":f"Catastral: {fmt_eur(catastral)} · Precio compra: {fmt_eur(precio)}",
                "accion":"Calcular 3% s/ MAX(precio compra, catastral) × % construcción"})

    if ibi == 0 and renta > 0:
        problemas.append({"tipo":"crit","titulo":"IBI no registrado",
            "desc":"Casilla 0106 a cero.","accion":"Solicitar recibo IBI 2024 al propietario"})

    if seguro == 0 and renta > 0:
        problemas.append({"tipo":"warn","titulo":"Seguro de hogar no registrado",
            "desc":"Puede ser deducible.","accion":"Verificar si el propietario tiene póliza"})

    if comunidad == 0 and renta > 0:
        problemas.append({"tipo":"warn","titulo":"Comunidad de propietarios a 0",
            "desc":"Revisar si tiene gastos de comunidad.","accion":"Confirmar con propietario"})

    gastos_total = ibi + seguro + comunidad*12 + amort
    if ingresos > 0 and gastos_total > ingresos * 0.70:
        problemas.append({"tipo":"warn","titulo":"Gastos > 70% de los ingresos",
            "desc":f"Gastos: {fmt_eur(gastos_total)} · Ingresos: {fmt_eur(ingresos)}",
            "accion":"Revisar si hay gastos duplicados o importes incorrectos"})

    es_larga = "larga" in tipo or "habitual" in tipo or tipo == ""
    # Alerta reducción 60% solo para personas físicas — no aplica a sociedades
    _tipo_ct_sem = ""
    try:
        import streamlit as _st_sem
        _tipo_ct_sem = _st_sem.session_state.get("fh_cartera",[])
    except: pass
    if es_larga:
        try:
            anyo = int(fecha_str[:4]) if fecha_str and len(fecha_str) >= 4 else 0
            if 0 < anyo <= 2023:
                problemas.append({"tipo":"warn","titulo":"Posible reducción 60% — confirmar",
                    "desc":f"Contrato desde {anyo}. Verificar condiciones Art. 23.2 LIRPF.",
                    "accion":"Confirmar fecha y condiciones antes de aplicar"})
        except: pass

    if any(p["tipo"] == "crit" for p in problemas):   estado = "cr"
    elif any(p["tipo"] == "warn" for p in problemas):  estado = "wn"
    else:                                               estado = "ok"

    return {"problemas": problemas, "estado": estado}

def calcular_modelo100_inmueble(row, df_mov):
    nombre    = str(row.get("nombre") or row.get("Nombre",""))
    renta     = _gv(row,"renta","Renta")
    dias      = int(_gv(row,"dias_arrendados_anio","Dias_Arrendados_Anio",d=365))
    factor    = min(dias,365)/365
    ingresos  = renta * 12 * factor
    intereses = _gv(row,"intereses_hipoteca","Intereses_Hipoteca") * factor
    ibi       = _gv(row,"ibi_anual","IBI_Anual") * factor
    comunidad = _gv(row,"comunidad","Comunidad") * 12 * factor
    seguro_h  = _gv(row,"seguro_anual","Seguro_Anual") * factor
    seguro_v  = _gv(row,"seguro_vida","Seguro_Vida") * factor
    ascensor  = _gv(row,"gasto_ascensor","Gasto_Ascensor") * factor
    com_seg   = comunidad + seguro_h + seguro_v + ascensor
    suministros = _gv(row,"servicios_suministros","Servicios_Suministros") * factor
    gastos_jur  = _gv(row,"gastos_juridicos","Gastos_Juridicos") * factor
    retenciones = _gv(row,"retenciones_irpf","Retenciones_IRPF")
    precio   = _gv(row,"precio_compra","Precio_Compra")
    imptos   = _gv(row,"impuestos_compra","Impuestos_Compra")
    gastos_c = _gv(row,"gastos_compra","Gastos_Compra")
    catastral= _gv(row,"valor_catastral","Valor_Catastral")
    pct_c    = _gv(row,"pct_construccion","Pct_Construccion",d=0.75)
    base_amort = max(precio+imptos+gastos_c, catastral)
    amort    = base_amort * pct_c * 0.03 * factor
    reparaciones = 0.0
    if not df_mov.empty:
        ca = "apartamento" if "apartamento" in df_mov.columns else "Apartamento"
        ct = "tipo" if "tipo" in df_mov.columns else "Tipo"
        cc = "categoria" if "categoria" in df_mov.columns else "Categoría"
        ci = "importe" if "importe" in df_mov.columns else "Importe"
        mask = ((df_mov.get(ca,pd.Series())==nombre) &
                (df_mov.get(ct,pd.Series())=="Gasto") &
                (df_mov.get(cc,pd.Series()).isin(["Mantenimiento","Reparación"])))
        reparaciones = float(df_mov[mask][ci].sum()) * factor if mask.any() else 0
    total_gastos = intereses+reparaciones+ibi+com_seg+suministros+gastos_jur+amort
    rend_neto    = ingresos - total_gastos
    tipo  = str(row.get("tipo_arrendamiento") or row.get("Tipo_Arrendamiento","")).lower()
    fecha = str(row.get("fecha_inicio_contrato") or row.get("Fecha_Inicio_Contrato","") or "")
    es_larga = "larga" in tipo or "habitual" in tipo or tipo == ""
    red_pct = 0
    if es_larga:
        try:
            anyo = int(fecha[:4]) if fecha and len(fecha)>=4 else 0
            red_pct = 60 if 0<anyo<=2023 else 50
        except: red_pct = 50
    reduccion = rend_neto * red_pct/100 if rend_neto > 0 else 0
    rend_final = rend_neto - reduccion
    return {
        "ingresos": round(ingresos,2), "intereses": round(intereses,2),
        "reparaciones": round(reparaciones,2), "ibi": round(ibi,2),
        "comunidad_seguros": round(com_seg,2), "suministros": round(suministros,2),
        "gastos_juridicos": round(gastos_jur,2), "amortizacion": round(amort,2),
        "total_gastos": round(total_gastos,2), "rend_neto": round(rend_neto,2),
        "red_pct": red_pct, "reduccion": round(reduccion,2),
        "rend_final": round(rend_final,2), "retenciones": round(retenciones,2), "dias": dias,
    }

def calcular_modelo100_global(df_inm, df_mov):
    if df_inm.empty: return {}
    total = {k:0 for k in ["ingresos","intereses","reparaciones","ibi","comunidad_seguros",
             "suministros","gastos_juridicos","amortizacion","total_gastos","rend_neto",
             "reduccion","rend_final","retenciones"]}
    for _, row in df_inm.iterrows():
        m = calcular_modelo100_inmueble(row, df_mov)
        for k in total: total[k] += m.get(k,0)
    return {k: round(v,2) for k,v in total.items()}

def calcular_alertas_cliente(df_inm, df_mov):
    alertas = []
    if df_inm.empty: return alertas
    for _, row in df_inm.iterrows():
        nombre = str(row.get("nombre") or row.get("Nombre",""))
        sem = calcular_semaforo_inmueble(row)
        for p in sem["problemas"]:
            alertas.append({**p, "inmueble": nombre, "categoria": "Fiscal"})
    return alertas

def calcular_modelo200_global(df_inm, df_mov):
    """Calcula IS para todos los inmuebles de un cliente sociedad patrimonial."""
    if df_inm.empty:
        return {"ingresos": 0, "total_gastos": 0, "amortizacion": 0,
                "resultado": 0, "cuota_is": 0, "retenciones": 0}
    total_ingresos = 0
    total_gastos   = 0
    total_amort    = 0
    total_retenc   = 0

    for _, row in df_inm.iterrows():
        modelo = calcular_modelo100_inmueble(row, df_mov)
        ingresos  = sf(modelo.get("0102", modelo.get("ingresos", 0)))
        gastos    = sf(modelo.get("0107", modelo.get("total_gastos", 0)))
        amort     = sf(modelo.get("0113", modelo.get("amortizacion", 0)))
        retenc    = sf(modelo.get("0153", modelo.get("retenciones", 0)))
        total_ingresos += ingresos
        total_gastos   += gastos
        total_amort    += amort
        total_retenc   += retenc

    # En IS no aplica reducción del 60%
    resultado  = round(total_ingresos - total_gastos, 2)
    cuota_is   = round(max(resultado * 0.25, 0), 2)
    diferencial = round(cuota_is - total_retenc, 2)

    return {
        "ingresos":     total_ingresos,
        "total_gastos": total_gastos,
        "amortizacion": total_amort,
        "resultado":    resultado,       # [399] Base imponible
        "cuota_is":     cuota_is,        # [562] Cuota IS 25%
        "retenciones":  total_retenc,    # [582]
        "diferencial":  diferencial,     # [599] A ingresar
    }


def construir_cartera(clientes_vinculados):
    cartera = []
    for acc in clientes_vinculados:
        pid = acc.get("propietario_id") or acc.get("user_id")
        nombre_raw = acc.get("nombre") or acc.get("nombre_propietario","")
        email_raw  = acc.get("email","")
        if nombre_raw and " " not in nombre_raw and "@" not in nombre_raw:
            nombre = nombre_raw
        elif email_raw:
            nombre = email_raw.split("@")[0].replace("."," ").title()
        else:
            nombre = nombre_raw or "Propietario"
        if not pid: continue
        df_inm   = get_inmuebles_propietario(pid)
        df_mov   = get_movimientos_propietario(pid)
        perfil   = get_perfil_propietario(pid)
        tipo     = perfil.get("tipo_cuenta", "particular")
        alertas  = calcular_alertas_cliente(df_inm, df_mov)
        modelo   = calcular_modelo100_global(df_inm, df_mov)
        modelo_is = calcular_modelo200_global(df_inm, df_mov) if tipo == "sociedad" else {}
        criticas = len([a for a in alertas if a["tipo"]=="crit"])
        medias   = len([a for a in alertas if a["tipo"]=="warn"])
        impacto  = sum(a.get("impacto",0) for a in alertas)
        estado   = "critico" if criticas>0 else "medio" if medias>0 else "ok"
        cartera.append({
            "id": pid, "nombre": nombre,
            "inmuebles": len(df_inm), "criticas": criticas, "medias": medias,
            "impacto": impacto, "estado": estado,
            "alertas": alertas, "df_inm": df_inm, "df_mov": df_mov,
            "modelo100": modelo, "modelo_is": modelo_is,
            "tipo_cuenta": tipo,
            "nombre_sociedad": perfil.get("nombre_sociedad", nombre),
            "cif_sociedad":    perfil.get("cif_sociedad", ""),
        })
    cartera.sort(key=lambda x:({"critico":0,"medio":1,"ok":2}[x["estado"]],-x["criticas"]))
    return cartera

# ── PERFIL ASESOR ────────────────────────────────────────────────
def pantalla_perfil():
    asesor      = st.session_state.get("fh_asesor", {})
    user_id     = st.session_state.get("fh_user_id", "")
    nombre      = asesor.get("nombre","")
    despacho    = asesor.get("despacho","")
    logo_bytes  = st.session_state.get("fh_logo_bytes")

    st.markdown("""<div style="margin-bottom:20px;">
      <div class="nc-page-label">Configuracion del despacho</div>
      <div class="nc-page-title">Perfil del asesor</div>
    </div>""", unsafe_allow_html=True)

    col_form, col_prev = st.columns([6, 4])

    with col_form:
        st.markdown('''<div style="font-size:13px;font-weight:700;color:#1e293b;
            border-left:3px solid #534AB7;padding-left:10px;margin-bottom:12px;">
            Logo del despacho</div>''', unsafe_allow_html=True)

        if logo_bytes:
            st.image(logo_bytes, width=160, caption="Logo activo")
            st.caption("Este logo aparecera en todos los informes PDF.")
            if st.button("Eliminar logo", key="btn_del_logo"):
                guardar_logo(user_id, b"", "png")
                st.session_state.pop("fh_logo_bytes", None)
                st.success("Logo eliminado.")
                st.rerun()
        else:
            st.markdown('''
            <div style="background:#FFF5F5;border-radius:10px;padding:12px 16px;
                        margin-bottom:12px;border:1px solid #FCA5A5;
                        font-size:12px;color:#DC2626;">
                Sin logo — los informes PDF mostraran el icono por defecto
            </div>''', unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Subir logo del despacho (PNG o JPG, max 2MB)",
            type=["png","jpg","jpeg"],
            key="logo_uploader",
            help="Recomendado: fondo transparente, min 200x200px")

        if uploaded is not None:
            file_bytes = uploaded.read()
            ext = uploaded.name.split(".")[-1].lower()
            if ext == "jpeg": ext = "jpg"
            if len(file_bytes) > 2 * 1024 * 1024:
                st.error("El archivo supera 2MB. Comprime la imagen antes de subir.")
            else:
                st.image(file_bytes, width=120, caption="Vista previa")
                if st.button("Guardar logo", key="btn_subir_logo", type="primary"):
                    with st.spinner("Guardando..."):
                        ok = guardar_logo(user_id, file_bytes, ext)
                    if ok:
                        st.session_state["fh_logo_bytes"] = file_bytes
                        st.success("Logo guardado. Aparecera en los proximos PDFs.")
                        st.rerun()
                    else:
                        st.error("Error al guardar. Verifica que la tabla "
                                 "user_profiles existe en Supabase con columna logo_b64.")

        st.markdown('''<div style="font-size:13px;font-weight:700;color:#1e293b;
            border-left:3px solid #534AB7;padding-left:10px;
            margin:20px 0 12px;">Datos del despacho</div>''',
            unsafe_allow_html=True)
        nuevo_nombre   = st.text_input("Nombre del asesor", value=nombre,
                                      key="perfil_nombre")
        nuevo_despacho = st.text_input("Nombre del despacho", value=despacho,
                                       key="perfil_despacho")
        nuevo_telefono = st.text_input("Teléfono", value=asesor.get("telefono",""),
                                       key="perfil_telefono")
        nuevo_nif      = st.text_input("NIF / CIF", value=asesor.get("nif",""),
                                       key="perfil_nif")
        if st.button("💾 Guardar cambios", key="btn_guardar_perfil", type="primary"):
            payload = {
                "nombre":   nuevo_nombre,
                "despacho": nuevo_despacho,
                "telefono": nuevo_telefono,
                "nif":      nuevo_nif,
            }
            r = requests.patch(
                f"{SUPABASE_URL}/rest/v1/asesores?user_id=eq.{user_id}",
                headers={**_h(), "Prefer": "return=representation"},
                json=payload
            )
            if r.status_code in [200, 201]:
                st.session_state["fh_asesor"] = {**asesor, **payload}
                st.success("✅ Perfil actualizado correctamente.")
                st.rerun()
            else:
                st.error(f"Error al guardar: {r.text}")

    with col_prev:
        st.markdown('''<div style="font-size:13px;font-weight:700;color:#1e293b;
            border-left:3px solid #534AB7;padding-left:10px;margin-bottom:12px;">
            Vista previa cabecera PDF</div>''', unsafe_allow_html=True)

        if logo_bytes:
            import base64 as _b64p
            _b64_str = _b64p.b64encode(logo_bytes).decode()
            logo_display = (f'<img src="data:image/png;base64,{_b64_str}"' +
                            ' style="height:40px;width:auto;object-fit:contain;border-radius:4px;">')
        else:
            logo_display = ('<div style="width:40px;height:40px;background:#534AB7;' +
                            'border-radius:6px;display:flex;align-items:center;' +
                            'justify-content:center;font-weight:900;color:white;' +
                            'font-size:14px;">FH</div>')

        from datetime import date as _dp
        st.markdown(f"""
        <div style="background:#FFF;border:1.5px solid #E2E8F0;border-radius:10px;
                    padding:14px 16px;">
            <div style="display:flex;align-items:center;gap:10px;
                        padding-bottom:10px;border-bottom:2px solid #534AB7;">
                {logo_display}
                <div style="flex:1;">
                    <div style="font-size:15px;font-weight:900;color:#534AB7;">
                        FiscalHub</div>
                    <div style="font-size:10px;color:#64748B;">
                        Informe Fiscal de Arrendamiento</div>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:9px;color:#64748B;">
                        {_dp.today().strftime('%d/%m/%Y')}</div>
                    <div style="font-size:9px;color:#64748B;">
                        Asesor: {nombre}</div>
                </div>
            </div>
            <div style="font-size:10px;color:#94A3B8;margin-top:8px;
                        text-align:center;">
                Asi aparecera la cabecera en los informes PDF</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("""
        <div style="background:#F0EEFF;border-radius:8px;padding:10px 12px;
                    margin-top:12px;font-size:11px;color:#534AB7;">
            <b>Como funciona:</b> El logo se guarda en Supabase y se carga
            automaticamente al iniciar sesion. Aparecera en todos los
            informes PDF que generes desde cualquier ficha de inmueble.
        </div>""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────
def render_sidebar():
    asesor   = st.session_state.get("fh_asesor", {})
    nombre   = asesor.get("nombre","Asesor")
    despacho = asesor.get("despacho","Despacho Fiscal")
    email    = asesor.get("email","")
    iniciales= "".join(p[0].upper() for p in nombre.split()[:2])
    dias     = days_to_irpf()
    pct      = max(0, min(100, int((90-dias)/90*100)))
    color    = "#DC2626" if dias<30 else "#D97706" if dias<60 else "#059669"
    user_id  = st.session_state.get("fh_user_id","")
    logo_bytes = st.session_state.get("fh_logo_bytes")

    # Comprimir espaciado Streamlit en sidebar
    st.sidebar.markdown("""<style>
    section[data-testid="stSidebar"] .stButton {margin-bottom:-8px !important;}
    section[data-testid="stSidebar"] .stButton button {padding:7px 12px !important;}
    section[data-testid="stSidebar"] .element-container {margin-bottom:0 !important;}
    section[data-testid="stSidebar"] .stFileUploader {margin-bottom:0 !important;}
    section[data-testid="stSidebar"] .stExpander {margin-bottom:0 !important;}
    </style>""", unsafe_allow_html=True)

    # ── Cabecera: FiscalHub + email ───────────────────────────────
    import base64 as _b64sb
    if logo_bytes:
        _sb_b64 = _b64sb.b64encode(logo_bytes).decode()
        logo_html = (
            f'<img src="data:image/png;base64,{_sb_b64}" '
            f'style="height:44px;width:auto;max-width:90px;object-fit:contain;'
            f'border-radius:6px;display:block;">'
        )
    else:
        logo_html = ""

    st.sidebar.markdown(f"""
    <div class="sb-brand" style="padding-bottom:8px;">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:2px;">
        <div class="sb-logo">NC</div>
        <div class="sb-wordmark">FiscalHub</div>
      </div>
      <div class="sb-tag" style="padding-left:0;">Portal asesoría fiscal</div>
      {f'<div style="font-size:10px;color:rgba(255,255,255,0.45);margin-top:3px;">{email}</div>' if email else ''}
    </div>
    <div class="sb-advisor" style="display:flex;align-items:center;justify-content:space-between;gap:8px;">
      <div style="display:flex;align-items:center;gap:8px;flex:1;min-width:0;">
        <div class="sb-avatar">{iniciales}</div>
        <div style="min-width:0;">
          <div class="sb-advisor-name">{nombre}</div>
          <div class="sb-advisor-desc">{despacho}</div>
        </div>
      </div>
      {f'<div style="flex-shrink:0;">{logo_html}</div>' if logo_html else ''}
    </div>""", unsafe_allow_html=True)

    # ── Logo del despacho — botón quitar o subir ─────────────────
    if logo_bytes:
        if st.sidebar.button("✕ Quitar logo", key="fh_quitar_logo",
                             use_container_width=True):
            st.session_state.pop("fh_logo_bytes", None)
            guardar_logo(user_id, b"", "png")
            st.rerun()
    else:
        with st.sidebar.expander("🖼️ Logo del despacho", expanded=False):
            logo_file = st.file_uploader("PNG o JPG", type=["png","jpg","jpeg"],
                key="fh_upload_logo", label_visibility="collapsed")
            if logo_file:
                ext = logo_file.name.split(".")[-1].lower()
                if ext == "jpeg": ext = "jpg"
                b = logo_file.read()
                ok = guardar_logo(user_id, b, ext)
                if ok:
                    st.session_state["fh_logo_bytes"] = b
                    st.success("Logo guardado")
                    st.rerun()

    for label, key in [("🗂 Cartera","cartera"),("⚠️ Alertas","alertas"),
                        ("📥 Exportar","exportar"),("🔗 Vincular","vincular"),
                        ("⚙️ Perfil","perfil")]:
        if st.sidebar.button(label, key=f"sb_{key}", use_container_width=True):
            for k in ["fh_cliente_sel","fh_inmueble_sel"]:
                st.session_state.pop(k, None)
            st.session_state.fh_menu = key
            st.rerun()

    st.sidebar.markdown("---")
    st.sidebar.markdown(f"""
    <div class="sb-irpf">
      <div class="sb-irpf-label">Cierre IRPF 2025</div>
      <div class="sb-irpf-num" style="color:{color};">{dias}
        <span style="font-size:14px;font-weight:600;margin-left:4px;opacity:0.7;">días</span>
      </div>
      <div class="sb-irpf-sub">30 jun · campaña 2025</div>
      <div class="sb-bar"><div class="sb-fill" style="width:{pct}%;background:{color};"></div></div>
      <div class="sb-bar-labels"><span>hoy</span><span>30 jun</span></div>
    </div>""", unsafe_allow_html=True)

    if st.sidebar.button("🚪 Cerrar sesión", use_container_width=True):
        for k in ["fh_logged","fh_user_id","fh_token","fh_asesor","fh_menu",
                  "fh_cliente_sel","fh_inmueble_sel","fh_cartera","fh_validaciones"]:
            st.session_state.pop(k, None)
        st.rerun()

# ── LOGIN ─────────────────────────────────────────────────────────
def pantalla_login():
    inject_global_css("ficahub")
    st.markdown("<div style='height:12vh;'></div>", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 1, 1])
    with col:
        st.markdown("""
        <div style="text-align:center;margin-bottom:22px;">
          <div style="display:inline-flex;align-items:center;gap:10px;margin-bottom:6px;">
            <div style="width:32px;height:32px;border:2px solid #534AB7;border-radius:6px;display:flex;align-items:center;justify-content:center;font-family:'Playfair Display',serif;font-size:13px;color:#534AB7;font-weight:700;">NC</div>
            <span style="font-family:'Playfair Display',serif;font-size:24px;color:#1e293b;font-weight:700;">FiscalHub</span>
          </div>
          <div style="font-size:9px;letter-spacing:0.18em;text-transform:uppercase;color:#94A3B8;margin-top:4px;">Portal Asesoría Fiscal · Nolasco Capital</div>
        </div>
        <div style="background:#FFFFFF;border:1px solid rgba(83,74,183,0.12);border-radius:12px;padding:24px 22px 20px;box-shadow:0 4px 24px rgba(83,74,183,0.08);">
        """, unsafe_allow_html=True)

        tab_li, tab_re = st.tabs(["Acceder", "Registrarse"])
        with tab_li:
            em = st.text_input("", key="li_em", placeholder="email@despacho.es", label_visibility="collapsed")
            pw = st.text_input("", type="password", key="li_pw", placeholder="Contraseña", label_visibility="collapsed")
            if st.button("Entrar →", key="li_btn", use_container_width=True, type="primary"):
                if em and pw:
                    with st.spinner("Verificando..."):
                        res = login_asesor(em, pw)
                    if res["ok"]:
                        uid = res["user_id"]
                        st.session_state.update({
                            "fh_logged":  True,
                            "fh_user_id": uid,
                            "fh_token":   res["token"],
                            "fh_asesor":  get_asesor_info(uid),
                            "fh_menu":    "cartera"})
                        # Cargar logo si existe
                        logo = leer_logo(uid)
                        if logo:
                            st.session_state["fh_logo_bytes"] = logo
                        st.rerun()
                    else: st.error(res.get("error","Credenciales incorrectas"))
                else: st.warning("Introduce email y contraseña")

        with tab_re:
            c1,c2 = st.columns(2)
            with c1: nm = st.text_input("",key="rg_nm",placeholder="Nombre completo",label_visibility="collapsed")
            with c2: ds = st.text_input("",key="rg_ds",placeholder="Despacho",label_visibility="collapsed")
            em_r = st.text_input("",key="rg_em",placeholder="email@despacho.es",label_visibility="collapsed")
            pw_r = st.text_input("",type="password",key="rg_pw",placeholder="Contraseña (mín. 8 car.)",label_visibility="collapsed")
            if st.button("Crear cuenta →",key="rg_btn",use_container_width=True,type="primary"):
                if all([nm,ds,em_r,pw_r]):
                    if len(pw_r)<8: st.error("Mínimo 8 caracteres")
                    else:
                        with st.spinner("Creando..."): res = registrar_asesor(em_r,pw_r,nm,ds)
                        if res["ok"]: st.success("✅ Cuenta creada. Revisa tu email y accede.")
                        else: st.error(res.get("error","Error"))
                else: st.warning("Completa todos los campos")
        st.markdown("</div>", unsafe_allow_html=True)

# ── CARTERA ───────────────────────────────────────────────────────
def pantalla_cartera():
    cartera = st.session_state.get("fh_cartera", [])
    if not cartera:
        st.markdown("""<div style="text-align:center;padding:60px 20px;">
          <div style="font-size:36px;margin-bottom:14px;">🔗</div>
          <div style="font-family:'DM Serif Display',serif;font-size:22px;color:#1E2A3A;margin-bottom:8px;">Sin clientes vinculados</div>
          <div style="font-size:13px;color:#5A6A7E;">Ve a Vincular e introduce el código que te dé el propietario desde Nolasco Capital.</div>
        </div>""", unsafe_allow_html=True)
        return

    total_inm  = sum(c["inmuebles"] for c in cartera)
    total_crit = sum(c["criticas"]  for c in cartera)
    total_imp  = sum(c["impacto"]   for c in cartera)
    n_crit = sum(1 for c in cartera if c["estado"]=="critico")
    n_med  = sum(1 for c in cartera if c["estado"]=="medio")
    n_ok   = sum(1 for c in cartera if c["estado"]=="ok")

    st.markdown(f"""<div style="margin-bottom:20px;">
      <div class="nc-page-label">Granada · Despacho fiscal</div>
      <div class="nc-page-title">Cartera de clientes</div>
      <div class="nc-page-sub">{len(cartera)} propietarios · {total_inm} inmuebles · campaña IRPF 2025</div>
    </div>""", unsafe_allow_html=True)

    render_kpi_row([
        {"label":"👥 Clientes",        "value":str(len(cartera)),
         "color":ACCENT_F,
         "subtitle":f"{n_crit} críticos · {n_med} revisar · {n_ok} OK"},
        {"label":"🏠 Inmuebles",       "value":str(total_inm),
         "color":ACCENT_F,             "subtitle":"Activos patrimoniales"},
        {"label":"🚨 Alertas críticas","value":str(total_crit),
         "color":RED,                  "subtitle":"Antes del 30 jun"},
        {"label":"💶 Impacto fiscal",  "value":fmt_eur(total_imp),
         "color":AMBER,                "subtitle":"Recuperable · cartera"},
    ])


    st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)

    cf1,cf2 = st.columns([3,1])
    with cf1: filtro = st.radio("",["Todos","Críticos","A revisar","OK"],horizontal=True,key="fh_filtro",label_visibility="collapsed")
    with cf2: busqueda = st.text_input("",placeholder="🔍 Buscar...",key="fh_busqueda",label_visibility="collapsed")

    rows = [c for c in cartera if
            (filtro=="Todos" or (filtro=="Críticos" and c["estado"]=="critico") or
             (filtro=="A revisar" and c["estado"]=="medio") or (filtro=="OK" and c["estado"]=="ok")) and
            (not busqueda or busqueda.lower() in c["nombre"].lower())]

    # Colores header por estado
    # Estado semántico para badge
    _ECOL  = {"critico":"#DC2626","medio":"#D97706","ok":"#059669"}
    _ELBL  = {"critico":"⚠ Crítico","medio":"◔ Revisar","ok":"✓ OK"}
    _BADGE = {"critico":"rgba(220,38,38,0.12)",
              "medio":"rgba(217,119,6,0.12)","ok":"rgba(5,150,105,0.12)"}

    def _cli_icon(nombre):
        n = nombre.lower()
        if any(x in n for x in ["bufete","abogad","juríd","legal"]): return "⚖️","Bufete"
        if any(x in n for x in ["inmo","piso","alquil"]): return "🏢","Inmobiliaria"
        if any(x in n for x in ["médic","clínic","salud"]): return "🏥","Clínica"
        if any(x in n for x in ["restaur","hostel","bar","café"]): return "🍽️","Hostelería"
        if any(x in n for x in ["sl","slu","s.l","s.a"]): return "🏛️","Empresa"
        return "👤","Particular"

    MAX_COLS = 4
    for fila_start in range(0, len(rows), MAX_COLS):
        fila_rows = rows[fila_start:fila_start+MAX_COLS]
        cols = st.columns(MAX_COLS)
        for col_idx, c in enumerate(fila_rows):
            estado   = c["estado"]
            hdr      = _color_cli(c["id"])  # color único por cliente
            txt      = _ECOL[estado]
            lbl      = _ELBL[estado]
            badge_bg = _BADGE[estado]
            icon, tipo = _cli_icon(c["nombre"])
            imp_v    = fmt_eur(abs(c["impacto"])) if c["impacto"] else "—"
            cr_col   = "#DC2626" if c["criticas"]>0 else "#64748B"
            med_col  = "#D97706" if c["medias"]>0   else "#64748B"
            _c_es_soc  = c.get("tipo_cuenta","particular") == "sociedad"
            modelo     = c.get("modelo100",{})
            modelo_is_ = c.get("modelo_is",{})
            if _c_es_soc:
                ingresos = fmt_eur(modelo_is_.get("ingresos", modelo.get("ingresos",0)))
                cuota_is_ = fmt_eur(modelo_is_.get("cuota_is", 0))
                _lbl_ing  = "📥 [318] Ingresos"
                _lbl_base = f"[562] Cuota IS 25%: {cuota_is_}"
                _badge_soc = ' <span style="background:rgba(5,150,105,0.2);color:#059669;font-size:9px;font-weight:700;padding:2px 6px;border-radius:4px;">🏢 IS</span>'
            else:
                ingresos   = fmt_eur(modelo.get("ingresos",0))
                base_imp_  = fmt_eur(modelo.get("rend_final",0))
                _lbl_ing   = "📥 0102 Ingresos"
                _lbl_base  = f"Base imp. est.: {base_imp_}"
                _badge_soc = ""
            with cols[col_idx]:
                hdr_html = (
                    f'<div style="background:{hdr};border-radius:12px 12px 0 0;' +
                    f'padding:14px 16px 12px;display:flex;align-items:center;gap:10px;margin-bottom:-1px;">' +
                    f'<span style="font-size:22px;">{icon}</span>' +
                    f'<div style="flex:1;min-width:0;">' +
                    f'<div style="font-size:18px;font-weight:800;color:#FFF;line-height:1.2;' +
                    f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{c["nombre"]}{_badge_soc}</div>' +
                    f'<div style="font-size:10px;color:rgba(255,255,255,0.65);margin-top:2px;">' +
                    f'<span style="font-size:13px;opacity:0.75;">{tipo} · {c["inmuebles"]} inmuebles</span></div></div>' +
                    f'<span style="background:rgba(255,255,255,0.15);color:#FFF;font-size:9px;' +
                    f'font-weight:700;padding:3px 8px;border-radius:6px;">{lbl}</span></div>'
                )
                body_html = (
                    '<div style="background:#FFF;border:2px solid #E2E8F0;border-top:none;' +
                    'border-radius:0 0 12px 12px;padding:14px 16px 12px;">' +
                    '<div style="display:flex;justify-content:space-between;margin-bottom:8px;' +
                    'padding-bottom:8px;border-bottom:1px solid #F1F5F9;">' +
                    f'<span style="font-size:14px;color:#94A3B8;font-weight:600;">Tipo</span>' +
                    f'<span style="font-size:14px;color:#1e293b;font-weight:700;">{tipo}</span></div>' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:5px;">' +
                    f'<span style="font-size:14px;color:#94A3B8;">{_lbl_ing}</span>' +
                    f'<span style="font-size:16px;font-weight:800;color:#059669;">{ingresos}</span></div>' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:5px;">' +
                    f'<span style="font-size:14px;color:#94A3B8;">🚨 Alertas críticas</span>' +
                    f'<span style="font-size:16px;font-weight:800;color:{cr_col};">{c["criticas"]}</span></div>' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:10px;">' +
                    f'<span style="font-size:14px;color:#94A3B8;">⚡ A revisar</span>' +
                    f'<span style="font-size:16px;font-weight:800;color:{med_col};">{c["medias"]}</span></div>' +
                    f'<div style="background:{badge_bg};border-radius:6px;padding:5px 10px;' +
                    f'text-align:center;font-size:14px;font-weight:700;color:{txt};">' +
                    f'{_lbl_base}</div></div>'
                )
                st.markdown(hdr_html + body_html, unsafe_allow_html=True)
                if st.button("→ Ver expediente completo",
                             key=f"cli_{c['id']}_{fila_start}_{col_idx}",
                             use_container_width=True):
                    st.session_state.fh_cliente_sel = c["id"]
                    st.session_state.fh_menu = "cliente"
                    st.rerun()
                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)



# ── PANTALLA CLIENTE ──────────────────────────────────────────────
def pantalla_cliente():
    cliente_id = st.session_state.get("fh_cliente_sel")
    cartera    = st.session_state.get("fh_cartera", [])
    cliente    = next((c for c in cartera if c["id"]==cliente_id), None)
    if not cliente: st.warning("Selecciona un cliente."); return

    df_inm     = cliente["df_inm"]
    df_mov     = cliente["df_mov"]
    modelo     = cliente["modelo100"]
    modelo_is  = cliente.get("modelo_is", {})
    nombre     = cliente["nombre"]
    tipo       = cliente.get("tipo_cuenta", "particular")
    es_sociedad = tipo == "sociedad"
    nom_soc    = cliente.get("nombre_sociedad", nombre)
    cif_soc    = cliente.get("cif_sociedad", "")
    vlds       = st.session_state.get("fh_validaciones", {}).get(cliente_id, {})

    if st.button("← Volver a cartera", key="cli_back"):
        st.session_state.fh_menu = "cartera"
        st.session_state.pop("fh_cliente_sel", None)
        st.session_state.pop("fh_inmueble_sel", None)
        st.rerun()

    if es_sociedad:
        st.markdown(f"""<div style="margin-bottom:14px;">
          <div class="nc-page-label">Impuesto de Sociedades · Modelo 200</div>
          <div class="nc-page-title">{nom_soc}</div>
          <div class="nc-page-sub">CIF {cif_soc} · {cliente["inmuebles"]} inmuebles · IS 25% · Ejercicio 2025</div>
        </div>""", unsafe_allow_html=True)
        # Badge sociedad
        st.markdown(
            '<div style="display:inline-block;background:#0d1a0d;border:1px solid #059669;'
            'border-radius:6px;padding:5px 12px;font-size:12px;color:#059669;'
            'margin-bottom:12px;">🏢 Sociedad Patrimonial · IS 25% · Sin reducción arrendamiento</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(f"""<div style="margin-bottom:14px;">
          <div class="nc-page-label">Revisión IRPF 2025</div>
          <div class="nc-page-title">{nombre}</div>
          <div class="nc-page-sub">{cliente["inmuebles"]} inmuebles · Campaña IRPF 2025</div>
        </div>""", unsafe_allow_html=True)

    _cc = _color_cli(cliente_id)

    if es_sociedad:
        resultado  = modelo_is.get("resultado", 0)
        cuota_is   = modelo_is.get("cuota_is", 0)
        diferencial = modelo_is.get("diferencial", 0)
        color_dif  = "#DC2626" if diferencial > 0 else "#059669"
        render_kpi_grid([
            {"label":"📥 [318] Ingresos arrend.",
             "value": fmt_eur(modelo_is.get("ingresos", 0)),
             "color": GREEN, "border_color": _cc, "subtitle": "Íntegros ejercicio"},
            {"label":"📤 [319] Gastos deducibles",
             "value": f"−{fmt_eur(modelo_is.get('total_gastos', 0))}",
             "color": RED,   "border_color": _cc, "subtitle": "IS — sin reducción 60%"},
            {"label":"⚖️ [399] Base imponible",
             "value": fmt_eur(resultado),
             "color": ACCENT_F, "border_color": _cc, "subtitle": "Resultado neto IS"},
            {"label":"🏛️ [562] Cuota IS 25%",
             "value": fmt_eur(cuota_is),
             "color": AMBER, "border_color": _cc, "subtitle": "Tipo general Art. 29 LIS"},
        ])
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
        render_kpi_grid([
            {"label":"💳 [582] Retenciones",
             "value": fmt_eur(modelo_is.get("retenciones", 0)),
             "color": GREY, "border_color": _cc, "subtitle": "Ingresos a cuenta"},
            {"label":"📋 [599] A ingresar",
             "value": fmt_eur(diferencial),
             "color": color_dif, "border_color": _cc,
             "subtitle": "⚠️ A pagar" if diferencial > 0 else "✅ A devolver"},
            {"label":"🏗️ [320] Amortización",
             "value": fmt_eur(modelo_is.get("amortizacion", 0)),
             "color": GREY, "border_color": _cc, "subtitle": "3% s/ valor construcción"},
            {"label":"📊 Tipo efectivo IS",
             "value": f"{round(cuota_is/resultado*100,1) if resultado>0 else 0:.1f}%",
             "color": ACCENT_F, "border_color": _cc, "subtitle": "Sobre base imponible"},
        ])
    else:
        render_kpi_grid([
            {"label":"📥 0102 Ingresos",
             "value": fmt_eur(modelo.get("ingresos", 0)),
             "color": GREEN, "border_color": _cc, "subtitle": "Rendimiento íntegro"},
            {"label":"📤 Gastos deducibles",
             "value": f"−{fmt_eur(modelo.get('total_gastos', 0))}",
             "color": RED,   "border_color": _cc, "subtitle": "Total deducible"},
            {"label":"⚖️ 0149 Rend. neto",
             "value": fmt_eur(modelo.get("rend_neto", 0)),
             "color": ACCENT_F, "border_color": _cc, "subtitle": "Antes de reducción"},
            {"label":"🧾 0156 Base imp. est.",
             "value": fmt_eur(modelo.get("rend_final", 0)),
             "color": AMBER, "border_color": _cc, "subtitle": "⚠️ Orientativa"},
        ])


    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

    # Verificar si todos validados
    col_n = "nombre" if "nombre" in df_inm.columns else "Nombre"
    nombres_inm = [str(r.get(col_n,"")) for _,r in df_inm.iterrows()] if not df_inm.empty else []
    todos_ok = all(
        vlds.get(nm,{}).get("estado") in ("ok","vl") or
        calcular_semaforo_inmueble(df_inm[df_inm[col_n]==nm].iloc[0])["estado"] == "ok"
        for nm in nombres_inm
    ) if nombres_inm else False

    if todos_ok:
        st.markdown("""<div class="nc-callout pos" style="margin-bottom:8px;">
          <strong>✅ Todos los inmuebles revisados</strong> — listo para el resumen global y exportar.
        </div>""", unsafe_allow_html=True)
        if st.button("📊 Resumen global → Exportar", type="primary", key="cli_global"):
            st.session_state.fh_menu = "resumen_global"
            st.rerun()

    st.markdown('<div class="nc-section">Inmuebles</div>', unsafe_allow_html=True)
    if df_inm.empty:
        st.info("Sin inmuebles registrados para este cliente."); return

    # Color del cliente — mismo para todos sus inmuebles
    _cc = _color_cli(cliente_id)

    # Grid de 4 columnas — mismo patrón que cards de cliente
    MAX_COLS = 4
    inm_list = list(df_inm.iterrows())
    for fila_start in range(0, len(inm_list), MAX_COLS):
        fila_rows = inm_list[fila_start:fila_start+MAX_COLS]
        cols = st.columns(MAX_COLS)
        for col_idx, (_, row) in enumerate(fila_rows):
            idx        = fila_start + col_idx
            nombre_inm = str(row.get(col_n,""))
            sem        = calcular_semaforo_inmueble(row)
            vld        = vlds.get(nombre_inm,{})
            vld_estado = vld.get("estado","")
            vld_manual = vld.get("manual", False)

            # Métricas
            renta     = _gv(row,"renta","Renta")
            ibi       = _gv(row,"ibi_anual","IBI_Anual")
            amort     = _gv(row,"amortizacion_fiscal","Amortizacion_Fiscal")
            seguro    = _gv(row,"seguro_anual","Seguro_Anual")
            comunidad = _gv(row,"comunidad","Comunidad")*12
            gastos    = ibi+amort+seguro+comunidad
            neto      = renta*12 - gastos
            tipo_arr  = str(row.get("tipo_arrendamiento") or row.get("Tipo_Arrendamiento","Larga Duración"))
            inquilino = str(row.get("inquilino") or row.get("Inquilino","—"))[:28]

            # Estado visual
            if vld_estado in ("ok","vl"):
                est_lbl = "✓ Validado" if vld_manual else "✓ Correcto"
                est_bg  = "rgba(5,150,105,0.10)"
                est_col = "#059669"
            elif sem["estado"] in ("cr","critico"):
                # Filtrar alerta reducción 60% para sociedades en semáforo
                _probs_sem = sem["problemas"]
                if es_sociedad:
                    _probs_sem = [p for p in _probs_sem
                                  if "60" not in p.get("titulo","") and
                                  "reduccion" not in p.get("titulo","").lower()]
                    sem = {**sem, "problemas": _probs_sem}
                n_cr    = len([p for p in sem["problemas"] if p["tipo"]=="crit"])
                est_lbl = f"⚠ {n_cr} crítico{'s' if n_cr>1 else ''}"
                est_bg  = "rgba(220,38,38,0.10)"
                est_col = "#DC2626"
            elif sem["estado"] in ("wn","advertencia"):
                n_wn    = len([p for p in sem["problemas"] if p["tipo"]=="warn"])
                est_lbl = f"◔ {n_wn} aviso{'s' if n_wn>1 else ''}"
                est_bg  = "rgba(217,119,6,0.10)"
                est_col = "#D97706"
            else:
                est_lbl = "✓ Correcto"
                est_bg  = "rgba(5,150,105,0.10)"
                est_col = "#059669"

            # Primera alerta
            alerta_txt = ""
            if sem["problemas"] and vld_estado not in ("ok","vl"):
                p0 = sem["problemas"][0]["titulo"]
                mas = f" +{len(sem['problemas'])-1} más" if len(sem["problemas"])>1 else ""
                alerta_txt = f"⚠ {p0}{mas}"
            elif vld_manual:
                alerta_txt = "✓ Validado manualmente"

            neto_col = "#059669" if neto >= 0 else "#DC2626"

            with cols[col_idx]:
                # Header con color del cliente
                hdr = (
                    f'<div style="background:{_cc};border-radius:12px 12px 0 0;' +
                    f'padding:12px 14px 10px;margin-bottom:-1px;">' +
                    f'<div style="font-size:15px;font-weight:800;color:#FFF;' +
                    f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">' +
                    f'{nombre_inm}</div>' +
                    f'<div style="font-size:11px;color:rgba(255,255,255,0.65);margin-top:2px;">' +
                    f'{inquilino} · {tipo_arr}</div>' +
                    f'</div>'
                )
                body = (
                    f'<div style="background:#FFF;border:2px solid #E2E8F0;' +
                    f'border-top:none;border-radius:0 0 12px 12px;' +
                    f'padding:12px 14px 10px;">' +
                    # alerta
                    (f'<div style="font-size:12px;color:{est_col};font-weight:600;' +
                     f'margin-bottom:8px;padding:4px 8px;background:{est_bg};' +
                     f'border-radius:6px;">{alerta_txt}</div>' if alerta_txt else
                     f'<div style="height:4px;"></div>') +
                    # métricas
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:5px;">' +
                    f'<span style="font-size:12px;color:#94A3B8;">📈 Renta/mes</span>' +
                    f'<span style="font-size:13px;font-weight:800;color:#059669;">{fmt_eur(renta)}</span></div>' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:5px;">' +
                    f'<span style="font-size:12px;color:#94A3B8;">📉 Gastos/año</span>' +
                    f'<span style="font-size:13px;font-weight:800;color:#DC2626;">−{fmt_eur(gastos)}</span></div>' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:8px;">' +
                    f'<span style="font-size:12px;color:#94A3B8;">⚖️ Neto/año</span>' +
                    f'<span style="font-size:13px;font-weight:800;color:{neto_col};">{fmt_eur(neto)}</span></div>' +
                    # badge estado
                    f'<div style="background:{est_bg};border-radius:6px;padding:4px 10px;' +
                    f'text-align:center;font-size:12px;font-weight:700;color:{est_col};">' +
                    f'{est_lbl}</div></div>'
                )
                st.markdown(hdr + body, unsafe_allow_html=True)

                # Botones — Revisar y Validar
                b1, b2 = st.columns(2)
                with b1:
                    if st.button(f"🔍 Revisar",
                                 key=f"rev_{cliente_id[:8]}_{idx}",
                                 use_container_width=True):
                        st.session_state.fh_inmueble_sel = nombre_inm
                        st.session_state.fh_menu = "ficha"
                        st.rerun()
                with b2:
                    if vld_manual:
                        if st.button("↩ Desvalidar",
                                     key=f"dvl_{cliente_id[:8]}_{idx}",
                                     use_container_width=True):
                            st.session_state.fh_validaciones[cliente_id].pop(nombre_inm, None)
                            st.rerun()
                    elif vld_estado not in ("ok","vl") and sem["estado"] not in ("ok",""):
                        if st.button("✅ Validar",
                                     key=f"vld_{cliente_id[:8]}_{idx}",
                                     use_container_width=True):
                            if "fh_validaciones" not in st.session_state:
                                st.session_state.fh_validaciones = {}
                            if cliente_id not in st.session_state.fh_validaciones:
                                st.session_state.fh_validaciones[cliente_id] = {}
                            st.session_state.fh_validaciones[cliente_id][nombre_inm] = {
                                "estado":"vl","manual":True,
                                "fecha":date.today().strftime("%d/%m/%Y")}
                            st.rerun()
                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ── FICHA INMUEBLE ────────────────────────────────────────────────
def pantalla_ficha_inmueble():
    cliente_id  = st.session_state.get("fh_cliente_sel")
    nombre_inm  = st.session_state.get("fh_inmueble_sel","")
    cartera     = st.session_state.get("fh_cartera", [])
    cliente     = next((c for c in cartera if c["id"]==cliente_id), None)
    if not cliente or not nombre_inm: st.warning("Selecciona un inmueble."); return

    df_inm = cliente["df_inm"]
    df_mov = cliente["df_mov"]
    col_n  = "nombre" if "nombre" in df_inm.columns else "Nombre"
    rows   = df_inm[df_inm[col_n]==nombre_inm]
    if rows.empty: st.warning(f"No se encontró: {nombre_inm}"); return
    row = rows.iloc[0]

    es_sociedad = cliente.get("tipo_cuenta","particular") == "sociedad"
    sem    = calcular_semaforo_inmueble(row)
    # Filtrar alerta reducción 60% para sociedades
    if es_sociedad:
        sem["problemas"] = [p for p in sem.get("problemas",[])
                            if "60" not in p.get("titulo","") and
                            "reduccion" not in p.get("titulo","").lower()]
    modelo = calcular_modelo100_inmueble(row, df_mov)
    vlds   = st.session_state.get("fh_validaciones",{}).get(cliente_id,{})

    c_back, c_vld = st.columns([3,1])
    with c_back:
        if st.button("← Volver al cliente", key="fic_back"):
            st.session_state.fh_menu = "cliente"
            st.session_state.pop("fh_inmueble_sel", None)
            st.session_state.pop("fh_pdf_export", None)
            st.rerun()
    with c_vld:
        if vlds.get(nombre_inm,{}).get("estado") != "vl" and sem["estado"] in ("cr","wn"):
            if st.button("✅ Validar manualmente", key="fic_vld", use_container_width=True):
                if "fh_validaciones" not in st.session_state: st.session_state.fh_validaciones = {}
                if cliente_id not in st.session_state.fh_validaciones: st.session_state.fh_validaciones[cliente_id] = {}
                st.session_state.fh_validaciones[cliente_id][nombre_inm] = {
                    "estado":"vl","manual":True,"fecha":date.today().strftime("%d/%m/%Y")}
                st.rerun()

    ROOF = {"Casa":("#6B2737","#8B3547"),"Despacho":("#185FA5","#1A6FBF"),
            "Garaje":("#4A5568","#5A6580"),"Apartamento":("#B8924A","#CFA55A")}
    def _rtype(tipo,nombre):
        t=(str(tipo)+" "+str(nombre)).lower()
        if any(x in t for x in ["despacho","oficina","salon"]): return "Despacho"
        if any(x in t for x in ["casa","chalet","abarqueros"]): return "Casa"
        if any(x in t for x in ["cochera","garaje"]): return "Garaje"
        return "Apartamento"
    rt   = _rtype(row.get("tipo_arrendamiento",""),nombre_inm)
    c1,c2= ROOF[rt]
    sem_color = {"cr":"var(--cr)","wn":"var(--wn)","ok":"var(--ok)"}[sem["estado"]]
    sem_label = {"cr":"🔴 Requiere acción","wn":"🟡 Revisar","ok":"🟢 Correcto"}[sem["estado"]]

    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#FFFFFF 0%,#F4F8FC 100%);border:0.5px solid rgba(74,122,181,0.2);border-radius:14px;overflow:hidden;margin-bottom:14px;box-shadow:0 2px 16px rgba(30,58,95,0.08);">
      <svg viewBox="0 0 600 52" style="display:block;width:100%;margin-bottom:-1px;" preserveAspectRatio="none" xmlns="http://www.w3.org/2000/svg">
        <defs><linearGradient id="rh" x1="0%" y1="0%" x2="100%" y2="100%">
          <stop offset="0%" stop-color="{c1}"/><stop offset="100%" stop-color="{c2}"/>
        </linearGradient></defs>
        <rect width="600" height="52" fill="url(#rh)"/>
        <text x="300" y="32" text-anchor="middle" font-family="DM Serif Display,serif" font-size="17" fill="white" opacity="0.95">{nombre_inm}</text>
        <text x="300" y="47" text-anchor="middle" font-family="IBM Plex Sans,sans-serif" font-size="10" fill="white" opacity="0.6">{rt} · {row.get("tipo_arrendamiento") or row.get("Tipo_Arrendamiento","Larga Duración")}</text>
      </svg>
      <div style="padding:10px 16px;display:flex;align-items:center;justify-content:space-between;">
        <div>
          <span style="font-size:13px;font-weight:500;color:#1E2A3A;">{row.get("inquilino") or row.get("Inquilino","Sin inquilino")}</span>
          <span style="font-size:11px;color:#8A9BB0;margin-left:10px;">CP {row.get("cp") or row.get("CP","—")}</span>
        </div>
        <span style="font-size:12px;font-weight:500;color:{sem_color};">{sem_label}</span>
      </div>
    </div>""", unsafe_allow_html=True)

    if sem["problemas"]:
        for p in sem["problemas"]:
            is_crit = p["tipo"] == "crit"
            dot_c   = "#DC2626" if is_crit else "#D97706"
            bg_c    = "#FFF5F5" if is_crit else "#FFFBEB"
            bd_c    = "#FCA5A5" if is_crit else "#FDE68A"
            st.markdown(f"""
            <div style="background:{bg_c};border:1px solid {bd_c};border-left:4px solid {dot_c};
                        border-radius:8px;padding:10px 14px;margin-bottom:6px;
                        display:flex;align-items:flex-start;gap:10px;">
              <div style="width:10px;height:10px;border-radius:50%;background:{dot_c};
                          flex-shrink:0;margin-top:3px;"></div>
              <div style="flex:1;min-width:0;">
                <div style="font-size:14px;font-weight:700;color:#1e293b;
                            margin-bottom:3px;">{p["titulo"]}</div>
                <div style="font-size:12px;color:#475569;
                            margin-bottom:3px;">{p["desc"]}</div>
                <div style="font-size:11px;color:{dot_c};font-weight:600;">
                    → {p["accion"]}</div>
              </div>
            </div>""", unsafe_allow_html=True)


    # ── LABORATORIO FISCAL — 3 CAPAS ─────────────────────────────

    def _simular_tabla(row_, d_):
        """Recalcula Modelo 100 según estados de los radio buttons de la tabla."""
        def _inc(dk, val):
            """Incluir valor si acc_{dk} == 'incluir', excluir si 'excluir'."""
            return val if d_.get(f"acc_{dk}", "incluir") == "incluir" else 0

        renta_m   = sf(row_.get("renta") or row_.get("Renta",0))
        ing       = renta_m * 12
        intereses = _inc("intereses", sf(row_.get("intereses_hipoteca",0)))
        ibi_v     = _inc("ibi",       sf(row_.get("ibi_anual",0)))
        com       = _inc("comunidad", sf(row_.get("comunidad",0))*12)
        seg       = _inc("seguro",    sf(row_.get("seguro_anual",0)))
        jur       = _inc("juridicos", sf(row_.get("gastos_juridicos_anual",0)))
        sumi      = _inc("suministros",sf(row_.get("suministros_anual",0)))

        # Reparaciones: gasto=100%, inversion=5%/año, excluir=0
        rep_tot   = sf(row_.get("reparaciones_anual",0))
        acc_rep   = d_.get("acc_rep","gasto")
        if acc_rep == "gasto":
            rep_ded = rep_tot
        elif acc_rep == "inversion":
            rep_ded = rep_tot * 0.05
        else:
            rep_ded = 0

        # Amortización 3%: aplicar=calculado, excluir=0
        _precio_c  = sf(row_.get("precio_compra",0))
        _catastral = sf(row_.get("valor_catastral",0))
        _pct_c     = sf(row_.get("porcentaje_construccion",0.7))
        _amort_c   = max(_precio_c, _catastral) * _pct_c * 0.03
        acc_amort  = d_.get("acc_amort","aplicar")
        amort_3    = _amort_c if (acc_amort=="aplicar" and _precio_c>0 and _catastral>0) else 0

        total_g = intereses + ibi_v + com + seg + jur + sumi + rep_ded + amort_3
        rend_n  = ing - total_g

        tipo_arr_ = str(row_.get("tipo_arrendamiento") or "").lower()
        es_temp   = any(x in tipo_arr_ for x in ["temporada","habitacion","turistic"])
        red_pct   = 0 if es_temp else d_.get("reduccion_pct", 50)
        reduccion = max(rend_n, 0) * red_pct / 100
        return dict(ingresos=round(ing,2), total_gastos=round(total_g,2),
                    rend_neto=round(rend_n,2), red_pct=red_pct,
                    reduccion=round(reduccion,2), rend_final=round(rend_n-reduccion,2))

    dec_key = f"dec_{cliente_id[:8]}_{nombre_inm[:10]}"
    if dec_key not in st.session_state:
        st.session_state[dec_key] = {}
    dec = st.session_state[dec_key]

    m         = modelo

    def _tipo_marginal(base: float) -> float:
        """Tipo marginal IRPF 2025 (estatal + autonómica Andalucía)
        calculado exclusivamente sobre la base inmobiliaria disponible.
        No incluye otras rentas del contribuyente."""
        if base <= 0:       return 0.19
        if base <= 12450:   return 0.19
        if base <= 20200:   return 0.24
        if base <= 35200:   return 0.30
        if base <= 60000:   return 0.37
        if base <= 300000:  return 0.45
        return 0.47

    TIPO_MARG = 0.30  # placeholder; se recalcula tras conocer base_orig
    from datetime import date as _date_
    hoy        = _date_.today()
    mes_actual = hoy.month
    meses_rest = 12 - mes_actual
    renta_mes  = sf(row.get("renta") or row.get("Renta", 0))
    ing_acum   = renta_mes * mes_actual
    ing_proy   = renta_mes * 12
    precio_c   = sf(row.get("precio_compra", 0))
    catastral  = sf(row.get("valor_catastral", 0))
    pct_const  = sf(row.get("porcentaje_construccion", 0.7))
    amort_calc = max(precio_c, catastral) * pct_const * 0.03
    tipo_arr_str = str(row.get("tipo_arrendamiento") or "").lower()
    es_temporada = any(x in tipo_arr_str for x in ["temporada","habitacion","turistic"])

    # Sin optimizar = sin ningún gasto deducido (peor caso posible)
    # Así el asesor solo puede mejorar la situación, nunca empeorarla
    _worst = {
        "acc_intereses":   "excluir",
        "acc_ibi":         "excluir",
        "acc_comunidad":   "excluir",
        "acc_seguro":      "excluir",
        "acc_suministros": "excluir",
        "acc_juridicos":   "excluir",
        "acc_rep":         "excluir",
        "acc_amort":       "excluir",
        "reduccion_pct":   50,
    }
    m_base     = _simular_tabla(row, _worst)
    base_orig  = m_base["rend_final"]
    TIPO_MARG  = _tipo_marginal(base_orig)
    cuota_orig = max(base_orig * TIPO_MARG, 0)

    st.markdown(
        '<div class="nc-section" style="margin-top:20px;">📋 Laboratorio Fiscal</div>',
        unsafe_allow_html=True)

    col_tabla, col_impacto = st.columns([6, 4])

    with col_tabla:
        # ── TABLA DE AUDITORÍA ESTRATÉGICA ───────────────────────
        # Cabecera
        st.markdown("""
        <div style="display:grid;grid-template-columns:32px 1fr 90px 1fr;
                    gap:0;background:#F1F5F9;border-radius:8px 8px 0 0;
                    padding:8px 12px;font-size:17px;font-weight:800;
                    color:#64748B;text-transform:uppercase;letter-spacing:0.08em;
                    margin-bottom:2px;">
            <div></div><div>Concepto</div>
            <div style="text-align:right;">Importe</div>
            <div style="padding-left:12px;">Acción</div>
        </div>""", unsafe_allow_html=True)

        def _sem_dot(color, tooltip=""):
            colors = {"verde":"#059669","amarillo":"#D97706","rojo":"#DC2626"}
            c = colors.get(color, "#94A3B8")
            return (f'<span title="{tooltip}" style="display:inline-block;' +
                    f'width:10px;height:10px;border-radius:50%;' +
                    f'background:{c};flex-shrink:0;"></span>')

        def _fila_header(titulo, color="#534AB7"):
            st.markdown(
                f'<div style="font-size:17px;font-weight:800;color:{color};' +
                f'text-transform:uppercase;letter-spacing:0.06em;' +
                f'padding:10px 12px 4px;border-bottom:1px solid #E2E8F0;' +
                f'margin-bottom:4px;">{titulo}</div>',
                unsafe_allow_html=True)

        # ── INGRESOS (solo informativo) ───────────────────────────
        _fila_header("📥 Ingresos")
        renta_mes_ = sf(row.get("renta") or row.get("Renta",0))
        st.markdown(
            f'<div style="display:grid;grid-template-columns:32px 1fr 90px 1fr;' +
            f'gap:0;padding:8px 12px;border-bottom:1px solid #F8FAFC;' +
            f'align-items:center;">' +
            f'<div>{_sem_dot("verde","Ingresos registrados")}</div>' +
            f'<div style="font-size:16px;color:#1e293b;font-weight:500;">{"[318]" if es_sociedad else "0102"} · Renta anual</div>' +
            f'<div style="font-size:17px;font-weight:800;color:#059669;text-align:right;">' +
            f'{fmt_eur(renta_mes_*12)}</div>' +
            f'<div style="padding-left:12px;font-size:17px;color:#94A3B8;">' +
            f'{fmt_eur(renta_mes_)}/mes</div></div>',
            unsafe_allow_html=True)

        # ── GASTOS FIJOS: Incluir / Excluir ──────────────────────
        _fila_header("📤 Gastos deducibles fijos")

        if es_sociedad:
            gastos_fijos = [
                ("[662]", "Intereses hipoteca (cta. 662)",
                 sf(row.get("intereses_hipoteca",0)), "intereses"),
                ("[631]", "IBI y tributos (cta. 631)",
                 sf(row.get("ibi_anual",0)), "ibi"),
                ("[629]", "Comunidad propietarios (cta. 629)",
                 sf(row.get("comunidad",0))*12, "comunidad"),
                ("[625]", "Seguro hogar + vida (cta. 625)",
                 sf(row.get("seguro_anual",0)), "seguro"),
                ("[628]", "Suministros (cta. 628)",
                 sf(row.get("suministros_anual",0)), "suministros"),
                ("[623]", "Gastos jurídicos (cta. 623)",
                 sf(row.get("gastos_juridicos_anual",0)), "juridicos"),
            ]
        else:
            gastos_fijos = [
                ("0105", "Intereses hipoteca",
                 sf(row.get("intereses_hipoteca",0)), "intereses"),
                ("0106", "IBI y tributos",
                 sf(row.get("ibi_anual",0)), "ibi"),
                ("0107", "Comunidad propietarios",
                 sf(row.get("comunidad",0))*12, "comunidad"),
                ("0110", "Seguro hogar + vida",
                 sf(row.get("seguro_anual",0)), "seguro"),
                ("0111", "Suministros",
                 sf(row.get("suministros_anual",0)), "suministros"),
                ("0112", "Gastos jurídicos",
                 sf(row.get("gastos_juridicos_anual",0)), "juridicos"),
            ]

        # Default: excluir siempre hasta que el asesor lo valide activamente
        def _default_acc(dk, valor):
            if f"acc_{dk}" in dec:
                return dec[f"acc_{dk}"]
            # Por defecto excluido — el asesor activa lo que valida
            return "excluir"

        for cas, label, valor, dk in gastos_fijos:
            sem_v   = "verde" if valor > 0 else "rojo"
            tip     = "Registrado — activa para incluir" if valor > 0 else "Sin registrar"
            default = _default_acc(dk, valor)
            dec[f"acc_{dk}"] = default

            col_s, col_l, col_v, col_tog = st.columns([0.4, 2.8, 1.0, 1.5])
            with col_s:
                st.markdown(_sem_dot(sem_v, tip), unsafe_allow_html=True)
            with col_l:
                st.markdown(
                    f'<div style="font-size:14px;color:#1e293b;line-height:2.2;">' +
                    f'<span style="color:#94A3B8;font-size:12px;">{cas} </span>' +
                    f'{label}</div>',
                    unsafe_allow_html=True)
            with col_v:
                color_v = "#059669" if valor > 0 else "#DC2626"
                st.markdown(
                    f'<div style="font-size:14px;font-weight:700;line-height:2.2;' +
                    f'color:{color_v};text-align:right;">' +
                    f'{fmt_eur(valor)}</div>',
                    unsafe_allow_html=True)
            with col_tog:
                tog = st.toggle(
                    "Incluir",
                    value=(default == "incluir"),
                    key=f"tog_{dk}_{dec_key}",
                    help=tip)
                dec[f"acc_{dk}"] = "incluir" if tog else "excluir"

            st.markdown("<div style='height:2px;border-bottom:1px solid #F1F5F9;"
                        "margin:0 0 2px'></div>", unsafe_allow_html=True)

        # ── REPARACIONES: Gasto / Inversión / Excluir ─────────────
        _fila_header("🔧 Mantenimiento e inversiones")
        rep_v   = sf(row.get("reparaciones_anual",0))
        rep_sem = "verde" if 0 < rep_v <= 2000 else ("amarillo" if rep_v > 2000 else "rojo")
        rep_tip = ("Importe alto — revisar clasificación" if rep_v > 2000
                   else ("Sin reparaciones registradas" if rep_v == 0
                         else "Registrado"))
        rep_def = dec.get("acc_rep", "gasto")

        col_s, col_l, col_v, col_a = st.columns([0.4, 2.5, 1.1, 2.2])
        with col_s:
            st.markdown(_sem_dot(rep_sem, rep_tip), unsafe_allow_html=True)
        with col_l:
            st.markdown(
                f'<div style="font-size:16px;color:#1e293b;">' +
                f'<span style="color:#94A3B8;font-size:17px;">0104 </span>' +
                f'Reparaciones</div>',
                unsafe_allow_html=True)
        with col_v:
            color_r = "#059669" if rep_v > 0 else "#94A3B8"
            st.markdown(
                f'<div style="font-size:17px;font-weight:700;' +
                f'color:{color_r};text-align:right;">' +
                f'{fmt_eur(rep_v)}</div>',
                unsafe_allow_html=True)
        with col_a:
            rep_opc = st.radio("",
                options=["gasto","inversion","excluir"],
                format_func=lambda x: {
                    "gasto":"🛠️ Gasto",
                    "inversion":"📈 Inversión",
                    "excluir":"❌ Excluir"}[x],
                index=["gasto","inversion","excluir"].index(rep_def),
                horizontal=True,
                key=f"acc_rep_{dec_key}",
                label_visibility="collapsed")
            dec["acc_rep"] = rep_opc

        if rep_opc == "inversion" and rep_v > 0:
            st.caption(f"→ Amortizable: {fmt_eur(rep_v*0.05)}/año × 20 años")

        # ── AMORTIZACIÓN 3% (cálculo automático) ──────────────────
        _lbl_amort_hdr = "⚡ Amortización construcción [681]" if es_sociedad else "⚡ Amortización construcción (0109)"
        _fila_header(_lbl_amort_hdr, color="#DC2626")
        amort_actual = sf(row.get("amortizacion_fiscal",0))
        datos_ok     = precio_c > 0 and catastral > 0

        if not datos_ok:
            sem_a = "rojo"
            tip_a = "Faltan precio compra o valor catastral"
        elif amort_actual == 0:
            sem_a = "amarillo"
            tip_a = "No aplicada — calcular ahora"
        else:
            sem_a = "verde"
            tip_a = "Amortización registrada"

        col_s, col_l, col_v, col_a = st.columns([0.4, 2.5, 1.1, 2.2])
        with col_s:
            st.markdown(_sem_dot(sem_a, tip_a), unsafe_allow_html=True)
        with col_l:
            _cas_amort_lbl = "[681]" if es_sociedad else "0109"
            st.markdown(
                f'<div style="font-size:16px;color:#1e293b;">' +
                f'<span style="color:#94A3B8;font-size:17px;">{_cas_amort_lbl} </span>' +
                'Amort. 3% construcción</div>',
                unsafe_allow_html=True)
            if datos_ok:
                st.markdown(
                    f'<div style="font-size:17px;color:#94A3B8;">' +
                    f'MAX({fmt_eur(precio_c)},{fmt_eur(catastral)}) × {pct_const*100:.0f}% × 3%' +
                    f' = {fmt_eur(amort_calc)}</div>',
                    unsafe_allow_html=True)
        with col_v:
            color_am = "#059669" if amort_actual > 0 else "#D97706" if datos_ok else "#DC2626"
            st.markdown(
                f'<div style="font-size:17px;font-weight:700;' +
                f'color:{color_am};text-align:right;">' +
                f'{fmt_eur(amort_actual)}</div>',
                unsafe_allow_html=True)
        with col_a:
            if not datos_ok:
                st.markdown(
                    '<div style="font-size:14px;color:#DC2626;padding-left:8px;">' +
                    '🔴 Faltan datos</div>',
                    unsafe_allow_html=True)
                dec["acc_amort"] = "excluir"
            else:
                am_def = dec.get("acc_amort", "aplicar")
                am_opc = st.radio("",
                    options=["aplicar","excluir"],
                    format_func=lambda x: (
                        f"✅ Aplicar ({fmt_eur(amort_calc)})"
                        if x=="aplicar" else "❌ Excluir"),
                    index=0 if am_def=="aplicar" else 1,
                    horizontal=True,
                    key=f"acc_amort_{dec_key}",
                    label_visibility="collapsed")
                dec["acc_amort"] = am_opc

        # ── REDUCCIÓN: solo persona física ──────────────────────────
        if not es_sociedad:
            _fila_header("📋 Reducción arrendamiento")
            red_pct_v = 0 if es_temporada else 50
            red_label = "Temporada/Habitaciones — 0%" if es_temporada else "Vivienda habitual — 50% (Art. 23.2 LIRPF)"
            st.markdown(
                f'<div style="display:grid;grid-template-columns:32px 1fr;' +
                f'gap:0;padding:8px 12px;align-items:center;">' +
                f'<div>{_sem_dot("verde" if not es_temporada else "amarillo")}</div>' +
                f'<div style="font-size:16px;color:#475569;">{red_label}</div></div>',
                unsafe_allow_html=True)
            dec["reduccion_pct"] = red_pct_v
        else:
            dec["reduccion_pct"] = 0
            st.markdown(
                '<div style="background:#0d1a0d;border:1px solid #059669;border-radius:8px;' +
                'padding:8px 14px;font-size:13px;color:#059669;margin:8px 12px;">' +
                '🏢 IS: Reducción del 60% no aplica a personas jurídicas (Art. 23.2 LIRPF).' +
                '</div>',
                unsafe_allow_html=True)

        if st.button("↺ Restablecer decisiones", key=f"rst_{dec_key}"):
            st.session_state.pop(dec_key, None)
            st.rerun()

        st.session_state[dec_key] = dec

    with col_impacto:
        m_opt      = _simular_tabla(row, st.session_state.get(dec_key,{}))
        base_opt   = m_opt["rend_final"]
        if es_sociedad:
            TIPO_MARG  = 0.25
            cuota_opt  = max(base_opt * 0.25, 0)
            cuota_orig = max(base_orig * 0.25, 0)
        else:
            TIPO_MARG  = _tipo_marginal(base_opt)
            cuota_opt  = max(base_opt * TIPO_MARG, 0)
        ahorro_c   = cuota_orig - cuota_opt
        ahorro_b   = base_orig  - base_opt
        color_ok   = "#059669" if ahorro_c >= 0 else "#DC2626"
        color_ko   = "#DC2626"
        _lbl_imp  = "🏛️ Impacto IS — Modelo 200" if es_sociedad else "⚖️ Impacto fiscal comparado"
        _lbl_sin  = "Sin gastos deducidos" if es_sociedad else "Sin optimizar"
        _lbl_con  = "Con gastos deducidos" if es_sociedad else "Con tu asesor"
        _lbl_base_s = "[399] Resultado neto" if es_sociedad else "Base imponible"
        _lbl_base_c = "[399] Resultado neto" if es_sociedad else "Base fin ejercicio"
        _lbl_aho  = "Ahorro IS (25%)" if es_sociedad else "Ahorro fiscal"
        _lbl_tipo = "Tipo fijo IS: 25% (Art. 29 LIS)" if es_sociedad else f"Tipo marginal aplicado: {int(TIPO_MARG*100)}%"
        _desc_tip = ("IS tipo general 25%. Sin reducción por arrendamiento de vivienda habitual. Estimación orientativa."
                    ) if es_sociedad else (
                     "Calculado por tramos IRPF 2025 (escala estatal + Andalucía). "
                     "El tipo real puede ser superior si existen rentas adicionales.")

        st.markdown(f"""
        <div style="background:#FFF;border-radius:12px;border:2px solid #94A3B8;
                    padding:16px;margin-bottom:12px;
                    box-shadow:0 4px 12px rgba(0,0,0,0.08);">
          <div style="font-size:10px;font-weight:800;color:#94A3B8;
                      text-transform:uppercase;letter-spacing:0.08em;margin-bottom:12px;">
              {_lbl_imp}</div>
          <div style="display:grid;grid-template-columns:1fr 20px 1fr;align-items:center;">
            <div style="background:#FFF5F5;border-radius:8px;padding:10px;">
              <div style="font-size:9px;font-weight:700;color:{color_ko};
                          text-transform:uppercase;margin-bottom:4px;">{_lbl_sin}</div>
              <div style="font-size:11px;color:#475569;">{_lbl_base_s}</div>
              <div style="font-size:1.3rem;font-weight:900;color:{color_ko};">
                  {fmt_eur(base_orig)}</div>
              <div style="font-size:11px;color:#475569;margin-top:6px;">Cuota est.</div>
              <div style="font-size:1.1rem;font-weight:800;color:{color_ko};">
                  {fmt_eur(cuota_orig)}</div>
            </div>
            <div style="text-align:center;font-size:18px;color:#94A3B8;">→</div>
            <div style="background:#F0FDF4;border-radius:8px;padding:10px;
                        border:1.5px solid {color_ok};">
              <div style="font-size:9px;font-weight:700;color:{color_ok};
                          text-transform:uppercase;margin-bottom:4px;">{_lbl_con}</div>
              <div style="font-size:11px;color:#475569;">{_lbl_base_c}</div>
              <div style="font-size:1.3rem;font-weight:900;color:{color_ok};">
                  {fmt_eur(base_opt)}</div>
              <div style="font-size:11px;color:#475569;margin-top:6px;">Cuota est.</div>
              <div style="font-size:1.1rem;font-weight:800;color:{color_ok};">
                  {fmt_eur(cuota_opt)}</div>
            </div>
          </div>
          <div style="margin-top:10px;padding:10px;background:{'#F0FDF4' if ahorro_c>=0 else '#FFF5F5'};
                      border-radius:8px;border:1.5px solid {color_ok};
                      display:flex;justify-content:space-between;align-items:center;">
            <div>
              <div style="font-size:9px;font-weight:700;color:#475569;
                          text-transform:uppercase;">💶 {_lbl_aho}</div>
              <div style="font-size:1.4rem;font-weight:900;color:{color_ok};">
                  {fmt_eur(ahorro_c)}</div>
            </div>
            <div style="text-align:right;">
              <div style="font-size:10px;color:#475569;">Base reducida</div>
              <div style="font-size:1rem;font-weight:800;color:{color_ok};">
                  {fmt_eur(ahorro_b)}</div>
              <div style="font-size:9px;color:#94A3B8;margin-top:2px;">
                  ⚠️ Estimación · Año en curso</div>
            </div>
          </div>
          <div style="margin-top:10px;padding:8px 10px;background:#F8F9FA;
                      border-radius:8px;border:1px solid #E2E8F0;">
            <div style="font-size:9px;font-weight:700;color:#534AB7;
                        text-transform:uppercase;margin-bottom:4px;">
                📊 {_lbl_tipo}</div>
            <div style="font-size:9px;color:#64748B;line-height:1.5;text-align:justify;">
                {_desc_tip}
            </div>
          </div>
        </div>
        <div style="background:#F8F9FA;border-radius:10px;padding:12px;
                    border:1px solid #E2E8F0;margin-bottom:12px;">
          <div style="font-size:9px;font-weight:800;color:#94A3B8;
                      text-transform:uppercase;margin-bottom:8px;">
              📅 Proyección a 31 dic {hoy.year}</div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;">
            <div><div style="font-size:10px;color:#94A3B8;">Acumulado a hoy</div>
                 <div style="font-size:1rem;font-weight:800;color:#059669;">
                     {fmt_eur(ing_acum)}</div></div>
            <div><div style="font-size:10px;color:#94A3B8;">Proyección anual</div>
                 <div style="font-size:1rem;font-weight:800;color:#1e293b;">
                     {fmt_eur(ing_proy)}</div></div>
            <div><div style="font-size:10px;color:#94A3B8;">Meses restantes</div>
                 <div style="font-size:1rem;font-weight:800;color:#534AB7;">
                     {meses_rest}</div></div>
            <div><div style="font-size:10px;color:#94A3B8;">Reducción aplicada</div>
                 <div style="font-size:1rem;font-weight:800;color:#534AB7;">
                     {m_opt["red_pct"]}%</div></div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # SABIO IA — solo al pulsar botón
        # Icono Sabio IA — círculo morado con emoji cerebro
        robot_img = (
            '<span style="display:inline-flex;align-items:center;'
            'justify-content:center;width:32px;height:32px;flex-shrink:0;'
            'border-radius:50%;background:#534AB7;margin-right:8px;'
            'font-size:18px;line-height:1;">🧠</span>')

        st.markdown(
            f'<div style="display:flex;align-items:center;margin-bottom:8px;">' +
            robot_img +
            f'<span style="font-size:13px;font-weight:700;color:#534AB7;">' +
            f'Sabio IA · Análisis Fiscal</span></div>',
            unsafe_allow_html=True)

        analisis_key = f"ia_{dec_key}_{int(ahorro_b*10)}"
        if st.button("🔬 Generar análisis fiscal",
                     key=f"btn_ia_{dec_key}",
                     use_container_width=True, type="primary"):
            # Leer estados actuales de la tabla
            _sem_local  = calcular_semaforo_inmueble(row)
            _alertas    = "; ".join([p.get("titulo","")
                          for p in _sem_local.get("problemas",[])])
            _acc_rep    = dec.get("acc_rep","gasto")
            _acc_amort  = dec.get("acc_amort","aplicar")
            _rep_v      = sf(row.get("reparaciones_anual",0))
            _amort_v    = amort_calc if _acc_amort=="aplicar" else 0
            # Gastos incluidos/excluidos
            _gastos_inc = []
            _gastos_exc = []
            for _dk, _lbl in [("intereses","Intereses"),("ibi","IBI"),
                               ("comunidad","Comunidad"),("seguro","Seguro"),
                               ("suministros","Suministros"),("juridicos","Jurídicos")]:
                if dec.get(f"acc_{_dk}","incluir") == "incluir":
                    _gastos_inc.append(_lbl)
                else:
                    _gastos_exc.append(_lbl)

            prompt = (
                f"Inmueble: {nombre_inm} · "
                f"{row.get('tipo_arrendamiento','Larga Duración')} · "
                f"Renta {fmt_eur(renta_mes)}/mes\n\n"
                f"SIN OPTIMIZAR: Base {fmt_eur(base_orig)} · "
                f"Cuota {fmt_eur(cuota_orig)}\n\n"
                f"DECISIONES DEL ASESOR:\n"
                f"  Gastos incluidos: {', '.join(_gastos_inc) or 'ninguno'}\n"
                f"  Gastos excluidos: {', '.join(_gastos_exc) or 'ninguno'}\n"
                f"  Reparaciones ({fmt_eur(_rep_v)}): {_acc_rep} — "
                f"{'deducción 100%' if _acc_rep=='gasto' else 'amort. 5%/año' if _acc_rep=='inversion' else 'excluido'}\n"
                f"  Amortización 3%: {'aplicada ' + fmt_eur(_amort_v) if _acc_amort=='aplicar' else 'excluida'} "
                f"(correcta: {fmt_eur(amort_calc)})\n"
                f"  Reducción: {m_opt['red_pct']}%\n\n"
                f"RESULTADO: Base fin ejercicio {fmt_eur(base_opt)} · "
                f"Cuota {fmt_eur(cuota_opt)} · "
                f"Ahorro fiscal {fmt_eur(ahorro_c)}\n\n"
                f"ALERTAS SEMÁFORO: {_alertas or 'ninguna'}\n\n"
                f"4 frases: (1) valoración de las decisiones tomadas, "
                f"(2) algo más que optimizar o corregir, "
                f"(3) riesgo ante inspección de Hacienda, "
                f"(4) acción concreta antes del 31/12."
            )
            system = (
                "Eres el Asesor Fiscal IA de FiscalHub. Hablas con un asesor "
                "fiscal profesional. Terminología fiscal española. "
                "Máximo 4 frases con euros reales. "
                "Distingue lo seguro de lo arriesgado."
            )
            from sabio_fiscal import _llamar_claude
            with st.spinner("Analizando..."):
                resultado = _llamar_claude(system, prompt, max_tokens=400)
            st.session_state[analisis_key] = resultado

        if st.session_state.get(analisis_key):
            st.markdown(
                f'<div style="background:#F0EEFF;border-radius:10px;'
                f'padding:14px 16px;margin-top:8px;'
                f'border-left:4px solid #534AB7;'
                f'font-size:13px;color:#1e293b;line-height:1.7;">'
                f'{st.session_state[analisis_key]}'
                f'</div>', unsafe_allow_html=True)
            cr1, cr2 = st.columns(2)
            with cr1:
                if st.button("↺ Regenerar", key=f"regen_{dec_key}"):
                    st.session_state.pop(analisis_key, None)
                    st.rerun()
            with cr2:
                if st.button("🗑 Borrar", key=f"del_ia_{dec_key}"):
                    st.session_state.pop(analisis_key, None)
                    st.rerun()

    st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)
    e1, e2, _ = st.columns([1, 1, 2])
    with e1:
        gen_pdf = st.button("📄 Generar PDF", key="fic_pdf",
                            use_container_width=True, type="primary")
    with e2:
        if st.button("← Volver", key="ficha_volver", use_container_width=True):
            st.session_state.fh_menu = "cliente"
            st.rerun()

    if gen_pdf:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                Table, TableStyle, HRFlowable, Image as RLImage, KeepTogether)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            import io, os, base64, tempfile
            from datetime import date as _d

            # ── Logo desde session_state (base64 en memoria) ──────
            _logo_el    = None
            _logo_bytes = st.session_state.get("fh_logo_bytes")
            if _logo_bytes:
                try:
                    import tempfile as _tf
                    with _tf.NamedTemporaryFile(
                            suffix=".png", delete=False) as _tmp:
                        _tmp.write(_logo_bytes)
                        _tmp_path = _tmp.name
                    _logo_el = RLImage(_tmp_path,
                                       width=1.2*cm, height=1.2*cm)
                except Exception:
                    _logo_el = None

            buf = io.BytesIO()
            W, H = A4
            doc = SimpleDocTemplate(buf, pagesize=A4,
                leftMargin=2*cm, rightMargin=2*cm,
                topMargin=1.8*cm, bottomMargin=2*cm)

            PURPLE = colors.HexColor("#534AB7")
            DARK   = colors.HexColor("#0F172A")
            GRAY   = colors.HexColor("#64748B")
            GREEN_ = colors.HexColor("#059669")
            RED_   = colors.HexColor("#DC2626")
            LGRAY  = colors.HexColor("#F1F5F9")
            LPURP  = colors.HexColor("#F0EEFF")
            WHITE  = colors.white

            st_ = getSampleStyleSheet()
            def _ps(name, size, color=DARK, bold=False, align=TA_LEFT,
                    space_before=0, space_after=4, leading=None):
                return ParagraphStyle(name, parent=st_["Normal"],
                    fontSize=size,
                    textColor=color,
                    fontName="Helvetica-Bold" if bold else "Helvetica",
                    alignment=align,
                    spaceBefore=space_before,
                    spaceAfter=space_after,
                    leading=leading or size*1.4)

            p_brand  = _ps("brand",  18, PURPLE, bold=True, space_after=2)
            p_sub    = _ps("sub",     9, GRAY,   space_after=2)
            p_meta   = _ps("meta",    8, GRAY,   align=TA_RIGHT, space_after=0)
            p_h2     = _ps("h2",     11, DARK,   bold=True, space_before=14, space_after=5)
            p_body   = _ps("body",    9, DARK,   space_after=3)
            p_small  = _ps("small",   7, GRAY,   space_after=0)
            p_alert  = _ps("alert",   8, DARK,   space_after=3)

            # ── Función helper: tabla estilo ─────────────────────
            def _tbl_style(header_bg=PURPLE, stripe=LGRAY):
                return TableStyle([
                    ("BACKGROUND",    (0,0),(-1,0),  header_bg),
                    ("TEXTCOLOR",     (0,0),(-1,0),  WHITE),
                    ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",      (0,0),(-1,-1),  8),
                    ("ROWBACKGROUNDS",(0,1),(-1,-1), [stripe, WHITE]),
                    ("GRID",          (0,0),(-1,-1),  0.3,
                     colors.HexColor("#E2E8F0")),
                    ("ALIGN",         (0,0),(-1,0),  "CENTER"),
                    ("ALIGN",         (1,1),(-1,-1), "RIGHT"),
                    ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
                    ("PADDING",       (0,0),(-1,-1),  5),
                    ("ROWHEIGHT",     (0,0),(-1,-1),  14),
                ])

            elems = []

            # ── CABECERA con logo ─────────────────────────────────
            _es_soc_pdf = cliente.get("tipo_cuenta","particular") == "sociedad"
            _nom_soc_pdf = cliente.get("nombre_sociedad", nombre_inm)
            _cif_soc_pdf = cliente.get("cif_sociedad", "")
            _titulo_pdf  = ("Informe IS — Modelo 200 · Sociedad Patrimonial"
                            if _es_soc_pdf else "Informe Fiscal de Arrendamiento")
            hdr_data = [[
                _logo_el or Paragraph("FH", p_brand),
                [Paragraph("FiscalHub", p_brand),
                 Paragraph(_titulo_pdf, p_sub)],
                [Paragraph(f"Generado: {_d.today().strftime('%d/%m/%Y')}", p_meta),
                 Paragraph(f"Asesor: {st.session_state.get('fh_nombre_asesor') or st.session_state.get('fh_asesor',{}).get('nombre','—')}", p_meta)]
            ]]
            hdr_tbl = Table(hdr_data, colWidths=[1.5*cm, 10*cm, 5*cm])
            hdr_tbl.setStyle(TableStyle([
                ("VALIGN",  (0,0),(-1,-1), "MIDDLE"),
                ("PADDING", (0,0),(-1,-1), 0),
                ("LINEBELOW",(0,0),(-1,0), 1.5, PURPLE),
            ]))
            elems.append(hdr_tbl)
            elems.append(Spacer(1, 10))

            # ── Datos del inmueble ────────────────────────────────
            inq_str = str(row.get("inquilino") or row.get("Inquilino","—"))
            tipo_str= str(row.get("tipo_arrendamiento","Larga Duracion"))
            # Datos adicionales del inmueble
            titular_str   = str(row.get("titular") or row.get("Titular","—"))
            nif_prop_str  = str(row.get("nif_propietario") or row.get("NIF_Propietario","—"))
            nif_inq_str   = str(row.get("nif_inquilino") or row.get("NIF_Inquilino","—"))
            ref_cat_str   = str(row.get("ref_catastral") or row.get("Ref_Catastral","—"))
            dir_str       = str(row.get("direccion") or row.get("Direccion","—"))
            f_inicio_str  = str(row.get("fecha_inicio_contrato","—") or "—")
            f_fin_str     = str(row.get("fecha_vencimiento_contrato","—") or "—")
            val_cat_str   = fmt_eur(float(row.get("valor_catastral") or 0))
            amort_str     = fmt_eur(float(row.get("amortizacion_fiscal") or 0))
            ibi_str       = fmt_eur(float(row.get("ibi_anual") or 0))
            seguro_str    = fmt_eur(float(row.get("seguro_anual") or 0))
            comunidad_str = fmt_eur(float(row.get("comunidad") or 0))

            # Datos del asesor
            _as = st.session_state.get("fh_asesor", {})
            asesor_nombre   = _as.get("nombre","—")
            asesor_despacho = _as.get("despacho","—")
            asesor_nif      = _as.get("nif","—")
            asesor_tel      = _as.get("telefono","—")

            inm_data = [[
                Paragraph("<b>Propietario / Titular</b>", p_body),
                Paragraph(titular_str, p_body),
                Paragraph("<b>NIF propietario</b>", p_body),
                Paragraph(nif_prop_str, p_body),
            ],[
                Paragraph("<b>Inmueble</b>", p_body),
                Paragraph(nombre_inm, p_body),
                Paragraph("<b>Dirección</b>", p_body),
                Paragraph(dir_str, p_body),
            ],[
                Paragraph("<b>Ref. Catastral</b>", p_body),
                Paragraph(ref_cat_str, p_body),
                Paragraph("<b>Valor catastral</b>", p_body),
                Paragraph(val_cat_str, p_body),
            ],[
                Paragraph("<b>Inquilino</b>", p_body),
                Paragraph(inq_str, p_body),
                Paragraph("<b>NIF inquilino</b>", p_body),
                Paragraph(nif_inq_str, p_body),
            ],[
                Paragraph("<b>Tipo contrato</b>", p_body),
                Paragraph(tipo_str, p_body),
                Paragraph("<b>Renta mensual</b>", p_body),
                Paragraph(fmt_eur(renta_mes), p_body),
            ],[
                Paragraph("<b>Inicio contrato</b>", p_body),
                Paragraph(f_inicio_str, p_body),
                Paragraph("<b>Vencimiento</b>", p_body),
                Paragraph(f_fin_str, p_body),
            ],[
                Paragraph("<b>IBI anual</b>", p_body),
                Paragraph(ibi_str, p_body),
                Paragraph("<b>Seguro anual</b>", p_body),
                Paragraph(seguro_str, p_body),
            ],[
                Paragraph("<b>Comunidad</b>", p_body),
                Paragraph(comunidad_str, p_body),
                Paragraph("<b>Amortización fiscal</b>", p_body),
                Paragraph(amort_str, p_body),
            ]]
            # Bloque asesor
            elems.append(Spacer(1, 6))
            asesor_data = [[
                Paragraph("<b>Asesor</b>", p_body),
                Paragraph(asesor_nombre, p_body),
                Paragraph("<b>Despacho</b>", p_body),
                Paragraph(asesor_despacho, p_body),
            ],[
                Paragraph("<b>NIF asesor</b>", p_body),
                Paragraph(asesor_nif, p_body),
                Paragraph("<b>Teléfono</b>", p_body),
                Paragraph(asesor_tel, p_body),
            ]]
            as_tbl = Table(asesor_data, colWidths=[3*cm,5.5*cm,3*cm,5*cm])
            as_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0),(0,-1), colors.HexColor("#EEF2FF")),
                ("BACKGROUND", (2,0),(2,-1), colors.HexColor("#EEF2FF")),
                ("FONTNAME",   (0,0),(0,-1), "Helvetica-Bold"),
                ("FONTNAME",   (2,0),(2,-1), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0),(-1,-1), 8),
                ("GRID",       (0,0),(-1,-1), 0.3, colors.HexColor("#E2E8F0")),
                ("PADDING",    (0,0),(-1,-1), 5),
                ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
            ]))
            elems.append(as_tbl)
            elems.append(Spacer(1, 8))

            # Banner sociedad patrimonial
            if _es_soc_pdf:
                GREEN_IS = colors.HexColor("#059669")
                soc_data = [[
                    Paragraph(f"<b>🏢 Sociedad Patrimonial: {_nom_soc_pdf}</b>", p_body),
                    Paragraph(f"CIF: {_cif_soc_pdf}", p_body),
                    Paragraph("Impuesto de Sociedades · Tipo 25%", p_body),
                    Paragraph("Modelo 200 · Sin reducción Art.23.2", p_body),
                ]]
                soc_tbl = Table(soc_data, colWidths=[5.5*cm, 3*cm, 4*cm, 4*cm])
                soc_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#0d1a0d")),
                    ("TEXTCOLOR",  (0,0),(-1,0), colors.HexColor("#059669")),
                    ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
                    ("FONTSIZE",   (0,0),(-1,0), 8),
                    ("GRID",       (0,0),(-1,-1), 0.3, colors.HexColor("#059669")),
                    ("PADDING",    (0,0),(-1,-1), 6),
                    ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
                ]))
                elems.append(soc_tbl)
                elems.append(Spacer(1, 6))

            inm_tbl = Table(inm_data, colWidths=[3*cm,5.5*cm,3*cm,5*cm])
            inm_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0),(0,-1), LGRAY),
                ("BACKGROUND", (2,0),(2,-1), LGRAY),
                ("FONTNAME",   (0,0),(0,-1), "Helvetica-Bold"),
                ("FONTNAME",   (2,0),(2,-1), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0),(-1,-1), 8),
                ("GRID",       (0,0),(-1,-1), 0.3,
                 colors.HexColor("#E2E8F0")),
                ("PADDING",    (0,0),(-1,-1), 5),
                ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
            ]))
            elems.append(inm_tbl)
            elems.append(Spacer(1, 10))

            # ── Comparativa fiscal ────────────────────────────────
            m_opt_pdf  = _simular_tabla(row, dec)
            base_o     = m_base["rend_final"]
            base_s     = m_opt_pdf["rend_final"]
            if _es_soc_pdf:
                cuota_o    = round(max(base_o * 0.25, 0), 2)
                cuota_s    = round(max(base_s * 0.25, 0), 2)
            else:
                cuota_o    = round(max(base_o * 0.30, 0), 2)
                cuota_s    = round(max(base_s * 0.30, 0), 2)
            ahorro_pdf = round(cuota_o - cuota_s, 2)

            if _es_soc_pdf:
                elems.append(Paragraph("Impacto IS — Modelo 200: Sin gastos vs. Con gastos deducidos", p_h2))
                cmp_data = [
                    ["Concepto", "Sin gastos", "Con gastos", "Diferencia"],
                    ["[318] Ingresos arrendamiento",
                     fmt_eur(m_base["ingresos"]),
                     fmt_eur(m_opt_pdf["ingresos"]), "—"],
                    ["[319] Gastos deducibles IS",
                     fmt_eur(m_base["total_gastos"]),
                     fmt_eur(m_opt_pdf["total_gastos"]),
                     fmt_eur(m_opt_pdf["total_gastos"]-m_base["total_gastos"])],
                    ["[320] Amortización (3% construcción)",
                     fmt_eur(m_base.get("amortizacion",0)),
                     fmt_eur(m_opt_pdf.get("amortizacion",0)), "—"],
                    ["[399] RESULTADO NETO (Base imponible)",
                     fmt_eur(base_o),
                     fmt_eur(base_s),
                     fmt_eur(base_s-base_o)],
                    ["[562] Cuota IS — tipo general 25%",
                     fmt_eur(cuota_o),
                     fmt_eur(cuota_s),
                     fmt_eur(cuota_s-cuota_o)],
                ]
                cmp_tbl = Table(cmp_data, colWidths=[6.5*cm,2.8*cm,3.2*cm,4*cm])
                cmp_style = _tbl_style(header_bg=colors.HexColor("#059669"))
                cmp_style.add("FONTNAME",   (0,4),(-1,4), "Helvetica-Bold")
                cmp_style.add("BACKGROUND", (0,4),(-1,4), colors.HexColor("#E6F9F1"))
                cmp_style.add("TEXTCOLOR",  (0,4),(-1,4), colors.HexColor("#059669"))
            else:
                elems.append(Paragraph("Impacto Fiscal: Sin optimizar vs. Con asesor", p_h2))
                cmp_data = [
                    ["Concepto", "Sin optimizar", "Con asesor", "Diferencia"],
                    ["Ingresos anuales (0102)",
                     fmt_eur(m_base["ingresos"]),
                     fmt_eur(m_opt_pdf["ingresos"]), "—"],
                    ["Gastos deducibles",
                     fmt_eur(m_base["total_gastos"]),
                     fmt_eur(m_opt_pdf["total_gastos"]),
                     fmt_eur(m_opt_pdf["total_gastos"]-m_base["total_gastos"])],
                    ["Rendimiento neto (0149)",
                     fmt_eur(m_base["rend_neto"]),
                     fmt_eur(m_opt_pdf["rend_neto"]), "—"],
                    [f"Reduccion {m_base['red_pct']}% / {m_opt_pdf['red_pct']}%",
                     fmt_eur(m_base["reduccion"]),
                     fmt_eur(m_opt_pdf["reduccion"]), "—"],
                    ["BASE IMPONIBLE (0156)",
                     fmt_eur(base_o),
                     fmt_eur(base_s),
                     fmt_eur(base_s-base_o)],
                    ["Cuota est. IRPF (30%)",
                     fmt_eur(cuota_o),
                     fmt_eur(cuota_s),
                     fmt_eur(cuota_s-cuota_o)],
                ]
                cmp_tbl = Table(cmp_data, colWidths=[6*cm,3*cm,3.5*cm,4*cm])
                cmp_style = _tbl_style()
                cmp_style.add("FONTNAME",   (0,5),(-1,5), "Helvetica-Bold")
                cmp_style.add("BACKGROUND", (0,5),(-1,5), LPURP)
                cmp_style.add("TEXTCOLOR",  (0,5),(-1,5), PURPLE)
            cmp_tbl.setStyle(cmp_style)
            elems.append(cmp_tbl)

            # Caja de ahorro destacada
            ahorro_color = GREEN_ if ahorro_pdf >= 0 else RED_
            ahorro_data  = [[
                Paragraph("AHORRO FISCAL CONSEGUIDO", _ps("ak",9,WHITE,bold=True)),
                Paragraph(fmt_eur(ahorro_pdf),
                          _ps("av",14,WHITE,bold=True,align=TA_RIGHT)),
            ]]
            ahorro_tbl = Table(ahorro_data, colWidths=[12*cm,4.5*cm])
            ahorro_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0),(-1,0), ahorro_color),
                ("VALIGN",     (0,0),(-1,0), "MIDDLE"),
                ("PADDING",    (0,0),(-1,0), 8),
                ("ROUNDEDCORNERS", [4,4,4,4]),
            ]))
            elems.append(Spacer(1,4))
            elems.append(ahorro_tbl)
            elems.append(Spacer(1,10))

            # ── Ratios IS (solo si es sociedad) ───────────────────
            if _es_soc_pdf:
                elems.append(Paragraph("Análisis Fiscal IS — Ratios Societarios", p_h2))
                elems.append(Spacer(1,4))

                # Calcular valores IS del inmueble
                _r_anual  = sf(row.get("renta_mensual", row.get("Renta", 0))) * 12
                _ibi_r    = sf(row.get("ibi_anual", row.get("IBI_Anual", 0)))
                _seg_r    = sf(row.get("seguro_anual", row.get("Seguro_Anual", 0)))
                _com_r    = sf(row.get("comunidad", row.get("Comunidad", 0))) * 12
                _int_r    = sf(row.get("intereses_hipoteca", row.get("Intereses_Hipoteca", 0)))
                _amort_r  = amort_calc if dec.get("acc_amort","aplicar")=="aplicar" else 0
                _rep_r    = sf(row.get("reparaciones_anual", 0))
                _gas_op   = _ibi_r + _seg_r + _com_r + _rep_r
                _gas_tot  = _gas_op + _int_r + _amort_r
                _rto_sin  = _r_anual
                _rto_con  = _r_anual - _gas_tot
                _is_sin   = round(max(_rto_sin * 0.25, 0), 2)
                _is_con   = round(max(_rto_con * 0.25, 0), 2)
                _neto_sin = round(_rto_sin - _is_sin, 2)
                _neto_con = round(_rto_con - _is_con, 2)
                _div_sin  = round(max(_neto_sin * 0.19, 0), 2)
                _div_con  = round(max(_neto_con * 0.19, 0), 2)
                _ahorro_r = round(_is_sin - _is_con, 2)
                _final_sin = round(_neto_sin - _div_sin, 2)
                _final_con = round(_neto_con - _div_con, 2)
                _precio_r  = sf(row.get("precio_compra", row.get("Precio_Compra", 0)))
                _roi_r     = round(_rto_con / _precio_r * 100, 1) if _precio_r > 0 else 0
                _ebitda_r  = _r_anual - _gas_op

                GREEN_IS = colors.HexColor("#059669")
                RED_IS   = colors.HexColor("#DC2626")
                AMBER_IS = colors.HexColor("#D97706")

                # Tabla doble imposición
                _di_data = [
                    ["Concepto", "Sin gastos deducidos", "Con gastos deducidos"],
                    ["[399] Resultado neto",
                     fmt_eur(_rto_sin), fmt_eur(_rto_con)],
                    ["[562] IS 25%",
                     f"−{fmt_eur(_is_sin)}", f"−{fmt_eur(_is_con)}"],
                    ["Neto tras IS",
                     fmt_eur(_neto_sin), fmt_eur(_neto_con)],
                    ["Retención dividendos (19%)",
                     f"−{fmt_eur(_div_sin)}", f"−{fmt_eur(_div_con)}"],
                    ["NETO FINAL AL SOCIO",
                     fmt_eur(_final_sin), fmt_eur(_final_con)],
                ]
                _di_tbl = Table(_di_data, colWidths=[6*cm, 4.5*cm, 4.5*cm])
                _di_style = TableStyle([
                    ("BACKGROUND",     (0,0),(-1,0),  GREEN_IS),
                    ("TEXTCOLOR",      (0,0),(-1,0),  colors.white),
                    ("FONTNAME",       (0,0),(-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",       (0,0),(-1,-1),  8),
                    ("ROWBACKGROUNDS", (0,1),(-1,-2), [colors.HexColor("#F0FDF4"), colors.white]),
                    ("BACKGROUND",     (0,-1),(-1,-1), colors.HexColor("#0d1a0d")),
                    ("TEXTCOLOR",      (0,-1),(-1,-1), GREEN_IS),
                    ("FONTNAME",       (0,-1),(-1,-1), "Helvetica-Bold"),
                    ("GRID",           (0,0),(-1,-1),  0.3, colors.HexColor("#E2E8F0")),
                    ("ALIGN",          (1,0),(-1,-1),  "RIGHT"),
                    ("PADDING",        (0,0),(-1,-1),  5),
                ])
                _di_tbl.setStyle(_di_style)
                elems.append(_di_tbl)
                elems.append(Spacer(1, 8))

                # KPIs IS del activo en fila
                _kpi_data = [
                    ["KPI", "Valor", "Referencia"],
                    ["EBITDA del activo",
                     fmt_eur(_ebitda_r),
                     "Renta − gastos operativos (excl. intereses y amort.)"],
                    ["IS estimado 25%",
                     f"−{fmt_eur(_is_con)}",
                     "Cuota IS orientativa con gastos deducidos"],
                    ["ROI del activo",
                     f"{_roi_r:.1f}%" if _precio_r > 0 else "Sin precio compra",
                     "Resultado neto / inversión total"],
                    ["Tipo efectivo total (IS + dividendo)",
                     "38.75%",
                     "25% IS + 19% dividendo — doble imposición"],
                    ["Ahorro fiscal por deducir gastos",
                     fmt_eur(_ahorro_r),
                     "Diferencia IS sin vs con gastos"],
                ]
                _kpi_tbl = Table(_kpi_data, colWidths=[5*cm, 3*cm, 7*cm])
                _kpi_style = TableStyle([
                    ("BACKGROUND",     (0,0),(-1,0),  colors.HexColor("#0F172A")),
                    ("TEXTCOLOR",      (0,0),(-1,0),  colors.white),
                    ("FONTNAME",       (0,0),(-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",       (0,0),(-1,-1),  8),
                    ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.HexColor("#F8FAFC"), colors.white]),
                    ("FONTNAME",       (0,1),(0,-1),  "Helvetica-Bold"),
                    ("TEXTCOLOR",      (1,1),(1,-1),  GREEN_IS),
                    ("FONTNAME",       (1,1),(1,-1),  "Helvetica-Bold"),
                    ("GRID",           (0,0),(-1,-1),  0.3, colors.HexColor("#E2E8F0")),
                    ("ALIGN",          (1,0),(1,-1),  "RIGHT"),
                    ("PADDING",        (0,0),(-1,-1),  5),
                ])
                _kpi_tbl.setStyle(_kpi_style)
                elems.append(_kpi_tbl)
                elems.append(Spacer(1, 6))

                # Nota doble imposición
                _nota_di = Paragraph(
                    "⚠️ Doble imposición: El beneficio tributa al 25% en IS y si se distribuye como dividendo "
                    "tributa de nuevo al 19% (tipo general ahorro IRPF). Tipo efectivo total: 38.75%. "
                    "Valorar si compensa reinvertir en la sociedad vs distribuir dividendos. "
                    "Datos orientativos — consultar con asesor fiscal.",
                    _ps("nota_di", 7, colors.HexColor("#854F0B"))
                )
                _nota_box = Table([[_nota_di]], colWidths=[15*cm])
                _nota_box.setStyle(TableStyle([
                    ("BACKGROUND", (0,0),(-1,-1), colors.HexColor("#FFF9E6")),
                    ("BOX",        (0,0),(-1,-1), 0.5, colors.HexColor("#D97706")),
                    ("PADDING",    (0,0),(-1,-1), 6),
                    ("ROUNDEDCORNERS", [4,4,4,4]),
                ]))
                elems.append(_nota_box)
                elems.append(Spacer(1, 10))

            # ── Decisiones del asesor ─────────────────────────────
            elems.append(Paragraph("Decisiones aplicadas por el asesor", p_h2))
            dec_data = [["Casilla","Concepto","Importe","Accion"]]
            if _es_soc_pdf:
                _gastos_pdf = [
                    ("intereses","[662]","Intereses hipoteca (cta. 662)"),
                    ("ibi",      "[631]","IBI y tributos (cta. 631)"),
                    ("comunidad","[629]","Comunidad propietarios (cta. 629)"),
                    ("seguro",   "[625]","Seguro hogar y vida (cta. 625)"),
                    ("suministros","[628]","Suministros (cta. 628)"),
                    ("juridicos","[623]","Gastos jurídicos (cta. 623)"),
                ]
            else:
                _gastos_pdf = [
                    ("intereses","0105","Intereses hipoteca"),
                    ("ibi",      "0106","IBI y tributos"),
                    ("comunidad","0107","Comunidad propietarios"),
                    ("seguro",   "0110","Seguro hogar y vida"),
                    ("suministros","0111","Suministros"),
                    ("juridicos","0112","Gastos juridicos"),
                ]
            for _dk, _cas, _lbl in _gastos_pdf:
                _val = {"intereses":sf(row.get("intereses_hipoteca",0)),
                        "ibi":sf(row.get("ibi_anual",0)),
                        "comunidad":sf(row.get("comunidad",0))*12,
                        "seguro":sf(row.get("seguro_anual",0)),
                        "suministros":sf(row.get("suministros_anual",0)),
                        "juridicos":sf(row.get("gastos_juridicos_anual",0)),
                       }.get(_dk,0)
                _acc = dec.get(f"acc_{_dk}","incluir")
                dec_data.append([_cas, _lbl, fmt_eur(_val),
                                  "Incluido" if _acc=="incluir" else "Excluido"])
            # Reparaciones
            _ra  = dec.get("acc_rep","gasto")
            _rv  = sf(row.get("reparaciones_anual",0))
            _cas_rep_pdf  = "[621]" if _es_soc_pdf else "0104"
            _lbl_rep_pdf  = "Reparaciones (cta. 621 PGC)" if _es_soc_pdf else "Reparaciones y mantenimiento"
            _cas_amort_p  = "[681]" if _es_soc_pdf else "0109"
            _lbl_amort_p  = "Amortizacion 3% (cta. 681 PGC)" if _es_soc_pdf else "Amortizacion 3% construccion"
            dec_data.append([_cas_rep_pdf, _lbl_rep_pdf, fmt_eur(_rv),
                {"gasto":"Gasto directo 100%",
                 "inversion":"Inversion 5%/anio",
                 "excluir":"Excluido"}.get(_ra,"—")])
            _aa  = dec.get("acc_amort","aplicar")
            dec_data.append([_cas_amort_p, _lbl_amort_p,
                fmt_eur(amort_calc) if _aa=="aplicar" else "0 EUR",
                "Aplicada (calculo auto)" if _aa=="aplicar" else "No aplicada"])

            dec_tbl = Table(dec_data, colWidths=[1.8*cm,6*cm,2.5*cm,6.2*cm])
            dec_tbl.setStyle(_tbl_style())
            elems.append(dec_tbl)
            elems.append(Spacer(1,10))

            # ── Alertas semaforo ──────────────────────────────────
            _sem_pdf = calcular_semaforo_inmueble(row)
            _probs   = _sem_pdf.get("problemas",[])
            # Filtrar alerta reducción 60% si es sociedad — no aplica a personas jurídicas
            if _es_soc_pdf:
                _probs = [p for p in _probs if "reducción 60%" not in p.get("titulo","").lower()
                          and "reduccion 60" not in p.get("titulo","").lower()]
            if _probs:
                elems.append(Paragraph("Alertas detectadas por el semaforo fiscal", p_h2))
                al_data = [["Nivel","Concepto","Descripcion"]]
                for p_ in _probs:
                    nivel = "CRITICO" if p_.get("tipo")=="crit" else "REVISION"
                    al_data.append([nivel,
                        p_.get("titulo",""),
                        p_.get("descripcion","") or "Verificar"])
                al_tbl = Table(al_data, colWidths=[2*cm,5*cm,9.5*cm])
                al_style = _tbl_style(header_bg=RED_)
                # Colorear celdas CRITICO
                for ri_, row_ in enumerate(al_data[1:],1):
                    if row_[0]=="CRITICO":
                        al_style.add("TEXTCOLOR",(0,ri_),(0,ri_),RED_)
                        al_style.add("FONTNAME", (0,ri_),(0,ri_),"Helvetica-Bold")
                    else:
                        al_style.add("TEXTCOLOR",(0,ri_),(0,ri_),
                                     colors.HexColor("#D97706"))
                al_tbl.setStyle(al_style)
                elems.append(al_tbl)
                elems.append(Spacer(1,10))

            # ── Analisis IA si existe ─────────────────────────────
            ia_txt = st.session_state.get(analisis_key)
            if ia_txt:
                elems.append(Paragraph("Analisis del Sabio IA", p_h2))
                ia_data = [[Paragraph(ia_txt, p_body)]]
                ia_tbl  = Table(ia_data, colWidths=[16.5*cm])
                ia_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0,0),(-1,-1), LPURP),
                    ("LINERIGHT",  (0,0),(0,-1),  3, PURPLE),
                    ("PADDING",    (0,0),(-1,-1),  8),
                ]))
                elems.append(ia_tbl)
                elems.append(Spacer(1,10))

            # ── Pie de pagina ─────────────────────────────────────
            elems.append(HRFlowable(width="100%", thickness=0.5,
                color=GRAY, spaceAfter=4))
            _nota_pie = (
                "Documento orientativo generado por FiscalHub. "
                "Sociedad Patrimonial | IS tipo general 25% (Art. 29 LIS). "
                "Sin reduccion por arrendamiento de vivienda habitual. "
                "Verificar con software oficial AEAT antes de presentar el Modelo 200."
            ) if _es_soc_pdf else (
                "Documento orientativo generado por FiscalHub. "
                "Cuotas IRPF estimadas al tipo marginal del 30%. "
                "Verificar con software oficial AEAT antes de presentar."
            )
            elems.append(Paragraph(_nota_pie, p_small))

            doc.build(elems)
            buf.seek(0)
            st.download_button(
                "Descargar PDF",
                data=buf,
                file_name=f"FiscalHub_{nombre_inm.replace(' ','_').replace('/','_')}.pdf",
                mime="application/pdf",
                key="fic_pdf_dl")

        except ImportError:
            st.error("ReportLab no instalado. Añade reportlab a requirements.txt")
        except Exception as e_pdf:
            st.error(f"Error PDF: {str(e_pdf)[:200]}")


def pantalla_resumen_global():
    cliente_id = st.session_state.get("fh_cliente_sel")
    cartera    = st.session_state.get("fh_cartera", [])
    cliente    = next((c for c in cartera if c["id"]==cliente_id), None)
    if not cliente: st.warning("Selecciona un cliente."); return

    df_inm = cliente["df_inm"]; df_mov = cliente["df_mov"]; nombre = cliente["nombre"]
    vlds   = st.session_state.get("fh_validaciones",{}).get(cliente_id,{})
    modelo = calcular_modelo100_global(df_inm, df_mov)
    modelo_is_gl = cliente.get("modelo_is", {})
    es_soc_gl  = cliente.get("tipo_cuenta","particular") == "sociedad"
    nom_soc_gl = cliente.get("nombre_sociedad", nombre)
    cif_soc_gl = cliente.get("cif_sociedad", "")
    col_n  = "nombre" if "nombre" in df_inm.columns else "Nombre"
    nombres= [str(r.get(col_n,"")) for _,r in df_inm.iterrows()] if not df_inm.empty else []
    n_manual = sum(1 for nm in nombres if vlds.get(nm,{}).get("manual",False))

    if st.button("← Volver al cliente", key="gl_back"):
        st.session_state.fh_menu = "cliente"; st.rerun()

    if es_soc_gl:
        st.markdown(f"""<div style="margin-bottom:14px;">
          <div class="nc-page-label">Resumen global IS · Modelo 200 · 2025</div>
          <div class="nc-page-title">{nom_soc_gl}</div>
          <div class="nc-page-sub">CIF {cif_soc_gl} · {len(nombres)} inmuebles · IS 25% · Modelo 200</div>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown(f"""<div style="margin-bottom:14px;">
          <div class="nc-page-label">Resumen global IRPF 2025</div>
          <div class="nc-page-title">{nombre}</div>
          <div class="nc-page-sub">{len(nombres)} inmuebles · Modelo 100 consolidado</div>
        </div>""", unsafe_allow_html=True)

    if n_manual > 0:
        st.markdown(f"""<div class="nc-callout warn">
          <strong>⚠️ {n_manual} inmueble{"s" if n_manual>1 else ""} con validación manual</strong> —
          Verificar antes de presentar a la AEAT.
        </div>""", unsafe_allow_html=True)

    _cc_gl = _color_cli(cliente_id)
    _logo_gl = st.session_state.get("fh_logo_bytes")
    _asesor_gl = st.session_state.get("fh_asesor", {})
    if _logo_gl:
        import base64 as _b64gl2
        _b64gl_kpi = _b64gl2.b64encode(_logo_gl).decode()
    else:
        _b64gl_kpi = None
    if es_soc_gl:
        _res_gl  = modelo_is_gl.get("resultado", 0)
        _cuota_gl = modelo_is_gl.get("cuota_is", 0)
        _dif_gl   = modelo_is_gl.get("diferencial", 0)
        render_kpi_grid([
            {"label":"📥 [318] Ingresos",
             "value":fmt_eur(modelo_is_gl.get("ingresos",0)),
             "color":GREEN,    "border_color":_cc_gl, "subtitle":f"{len(nombres)} inmuebles"},
            {"label":"📤 [319] Gastos IS",
             "value":f"−{fmt_eur(modelo_is_gl.get('total_gastos',0))}",
             "color":RED,      "border_color":_cc_gl, "subtitle":"Sin reducción 60%"},
            {"label":"⚖️ [399] Base imponible",
             "value":fmt_eur(_res_gl),
             "color":ACCENT_F, "border_color":_cc_gl, "subtitle":"Resultado neto IS"},
            {"label":"🏛️ [562] Cuota IS 25%",
             "value":fmt_eur(_cuota_gl),
             "color":AMBER,    "border_color":_cc_gl, "subtitle":"⚠️ Orientativa"},
        ],
        logo_b64=_b64gl_kpi, despacho=_asesor_gl.get("despacho",""),
        asesor=_asesor_gl.get("nombre",""))
    else:
        render_kpi_grid([
            {"label":"📥 0102 Ingresos",
             "value":fmt_eur(modelo.get("ingresos",0)),
             "color":GREEN,    "border_color":_cc_gl, "subtitle":f"{len(nombres)} inmuebles"},
            {"label":"📤 Gastos deducibles",
             "value":f"−{fmt_eur(modelo.get('total_gastos',0))}",
             "color":RED,      "border_color":_cc_gl, "subtitle":"Total deducible"},
            {"label":"⚖️ 0149 Rend. neto",
             "value":fmt_eur(modelo.get("rend_neto",0)),
             "color":ACCENT_F, "border_color":_cc_gl, "subtitle":"Antes de reducción"},
            {"label":"🧾 0156 Base imp.",
             "value":fmt_eur(modelo.get("rend_final",0)),
             "color":AMBER,    "border_color":_cc_gl, "subtitle":"⚠️ Orientativa"},
        ],
    logo_b64=_b64gl_kpi,
    despacho=_asesor_gl.get("despacho",""),
    asesor=_asesor_gl.get("nombre",""))

    st.markdown('<div class="nc-section">Desglose por inmueble</div>', unsafe_allow_html=True)
    if not df_inm.empty:
        # Grid 3 columnas — mismo patrón de cards que pantalla_cliente
        cols_grid = st.columns(3)
        for idx_, (_, row_g) in enumerate(df_inm.iterrows()):
            nm_g   = str(row_g.get(col_n,""))
            m_g    = calcular_modelo100_inmueble(row_g, df_mov)
            manual = vlds.get(nm_g,{}).get("manual",False)
            tipo_g = str(row_g.get("tipo_arrendamiento") or
                         row_g.get("Tipo_Arrendamiento","Larga Duración"))
            inq_g  = str(row_g.get("inquilino") or row_g.get("Inquilino","—"))
            hdr_c  = _color_cli(cliente_id)
            badge  = ("✎ Manual" if manual else "✓ Auto")
            badge_c= ("#D97706" if manual else "#059669")
            rend_c = "#059669" if m_g["rend_final"] >= 0 else "#DC2626"

            hdr_html = f"""
            <div style="background:{hdr_c};border-radius:10px 10px 0 0;
                        padding:12px 14px 10px;">
              <div style="font-size:15px;font-weight:800;color:#FFF;
                          line-height:1.2;margin-bottom:4px;">{nm_g}</div>
              <div style="font-size:11px;color:rgba(255,255,255,0.8);">
                {inq_g[:22]} · {tipo_g}</div>
              <div style="margin-top:6px;">
                <span style="background:rgba(255,255,255,0.18);color:#FFF;
                             font-size:10px;font-weight:700;padding:2px 8px;
                             border-radius:12px;">{badge}</span>
              </div>
            </div>"""
            body_html = f"""
            <div style="background:#FFF;border:1px solid #E2E8F0;
                        border-top:none;border-radius:0 0 10px 10px;
                        padding:12px 14px;margin-bottom:4px;">
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">
                <div>
                  <div style="font-size:10px;color:#94A3B8;margin-bottom:2px;">
                    0102 Ingresos</div>
                  <div style="font-size:15px;font-weight:800;color:#059669;">
                    {fmt_eur(m_g["ingresos"])}</div>
                </div>
                <div>
                  <div style="font-size:10px;color:#94A3B8;margin-bottom:2px;">
                    Gastos totales</div>
                  <div style="font-size:15px;font-weight:800;color:#DC2626;">
                    −{fmt_eur(m_g["total_gastos"])}</div>
                </div>
                <div>
                  <div style="font-size:10px;color:#94A3B8;margin-bottom:2px;">
                    0149 Rend. neto</div>
                  <div style="font-size:14px;font-weight:700;color:#534AB7;">
                    {fmt_eur(m_g["rend_neto"])}</div>
                </div>
                <div>
                  <div style="font-size:10px;color:#94A3B8;margin-bottom:2px;">
                    Reducción {m_g["red_pct"]}%</div>
                  <div style="font-size:14px;font-weight:700;color:#475569;">
                    −{fmt_eur(m_g["reduccion"])}</div>
                </div>
              </div>
              <div style="margin-top:10px;padding-top:8px;
                          border-top:1px solid #F1F5F9;
                          display:flex;justify-content:space-between;
                          align-items:center;">
                <div>
                  <div style="font-size:10px;color:#94A3B8;">0156 Base imp. est.</div>
                  <div style="font-size:1.3rem;font-weight:900;color:{rend_c};">
                    {fmt_eur(m_g["rend_final"])}</div>
                </div>
              </div>
            </div>"""
            with cols_grid[idx_ % 3]:
                st.markdown(hdr_html + body_html, unsafe_allow_html=True)


    st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)
    asesor = st.session_state.get("fh_asesor",{})
    nombre_asesor = asesor.get("despacho", asesor.get("nombre",""))
    e1,e2 = st.columns(2)
    with e1:
        if st.button("📄 Exportar PDF completo", type="primary", use_container_width=True, key="gl_pdf"):
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                    Table, TableStyle, HRFlowable, Image as _RLImg2)
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_RIGHT, TA_CENTER
                import io, tempfile as _tf2
                from datetime import date as _d2

                _buf = io.BytesIO()
                _doc = SimpleDocTemplate(_buf, pagesize=A4,
                    leftMargin=2*cm, rightMargin=2*cm,
                    topMargin=1.8*cm, bottomMargin=2*cm)
                _s   = getSampleStyleSheet()
                _PU  = colors.HexColor("#534AB7")
                _DK  = colors.HexColor("#0F172A")
                _GR  = colors.HexColor("#64748B")
                _LG  = colors.HexColor("#F1F5F9")
                _LP  = colors.HexColor("#F0EEFF")

                def _ps(name, size, color=_DK, bold=False, align=0, sa=4):
                    from reportlab.lib.enums import TA_LEFT
                    return ParagraphStyle(name, parent=_s["Normal"],
                        fontSize=size, textColor=color,
                        fontName="Helvetica-Bold" if bold else "Helvetica",
                        alignment=align, spaceAfter=sa, leading=size*1.4)

                p_h1   = _ps("g_h1", 18, _PU, bold=True, sa=2)
                p_sub  = _ps("g_sub", 9, _GR, sa=0)
                p_rt   = _ps("g_rt",  9, _GR, align=2)
                p_h2   = _ps("g_h2", 11, _DK, bold=True, sa=4)
                p_bd   = _ps("g_bd",  9, _DK)
                p_sm   = _ps("g_sm",  7, _GR, sa=0)

                # ── Logo ──────────────────────────────────────────
                _g_logo = None
                _g_logo_bytes = st.session_state.get("fh_logo_bytes")
                if _g_logo_bytes:
                    try:
                        with _tf2.NamedTemporaryFile(suffix=".png", delete=False) as _t2:
                            _t2.write(_g_logo_bytes)
                            _tp = _t2.name
                        _g_logo = _RLImg2(_tp, width=1.4*cm, height=1.4*cm)
                    except Exception:
                        _g_logo = None

                _el = []

                # ── Cabecera ──────────────────────────────────────
                # Columna central: dos párrafos como lista
                _titulo_gl = "Resumen Global IS — Modelo 200" if es_soc_gl else "Resumen Global IRPF"
                _mid = [Paragraph("FiscalHub", p_h1),
                        Paragraph(_titulo_gl, p_sub)]
                _right = [Paragraph(f"Generado: {_d2.today().strftime('%d/%m/%Y')}", p_rt),
                          Paragraph(f"Asesor: {nombre_asesor or '—'}", p_rt)]

                _hdr = Table(
                    [[_g_logo or Paragraph("FH", p_h1), _mid, _right]],
                    colWidths=[1.6*cm, 10*cm, 4.9*cm])
                _hdr.setStyle(TableStyle([
                    ("VALIGN",    (0,0),(-1,-1), "MIDDLE"),
                    ("PADDING",   (0,0),(-1,-1), 0),
                    ("TOPPADDING",(0,0),(-1,-1), 0),
                    ("BOTTOMPADDING",(0,0),(-1,-1), 6),
                    ("LINEBELOW", (0,0),(-1,0), 1.5, _PU),
                ]))
                _el.append(_hdr)
                _el.append(Spacer(1, 8))

                # ── Datos cliente ─────────────────────────────────
                _cld = [
                    [Paragraph("<b>Cliente</b>",  p_bd), Paragraph(nombre, p_bd),
                     Paragraph("<b>Inmuebles</b>",p_bd), Paragraph(str(len(nombres)), p_bd)],
                    [Paragraph("<b>Ejercicio</b>",p_bd), Paragraph("2025", p_bd),
                     Paragraph("<b>Asesor</b>",  p_bd), Paragraph(nombre_asesor or "—", p_bd)],
                ]
                _clt = Table(_cld, colWidths=[3*cm, 5.5*cm, 3*cm, 5*cm])
                _clt.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(0,-1),_LG),
                    ("BACKGROUND",(2,0),(2,-1),_LG),
                    ("FONTNAME",  (0,0),(0,-1),"Helvetica-Bold"),
                    ("FONTNAME",  (2,0),(2,-1),"Helvetica-Bold"),
                    ("FONTSIZE",  (0,0),(-1,-1),8),
                    ("GRID",      (0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
                    ("PADDING",   (0,0),(-1,-1),5),
                    ("VALIGN",    (0,0),(-1,-1),"MIDDLE"),
                ]))
                _el.append(_clt)
                _el.append(Spacer(1, 10))

                # ── Tabla consolidada IS o IRPF ──────────────────
                if es_soc_gl:
                    _el.append(Paragraph("Modelo 200 consolidado — IS Sociedad Patrimonial", p_h2))
                    _mis_pdf = cliente.get("modelo_is", {})
                    _res_pdf = _mis_pdf.get("resultado", 0)
                    _cuota_pdf = _mis_pdf.get("cuota_is", 0)
                    _retenc_pdf = _mis_pdf.get("retenciones", 0)
                    _dif_pdf = _mis_pdf.get("diferencial", 0)
                    _gr = [
                        ["Casilla", "Concepto", "Importe"],
                        ["[318]",    "Ingresos arrendamiento",
                         fmt_eur(_mis_pdf.get("ingresos",0))],
                        ["[319]",    "Gastos deducibles IS",
                         f"−{fmt_eur(_mis_pdf.get('total_gastos',0))}"],
                        ["[320]",    "Amortizacion 3% construccion",
                         fmt_eur(_mis_pdf.get("amortizacion",0))],
                        ["[399]",    "RESULTADO NETO (Base imponible)",
                         fmt_eur(_res_pdf)],
                        ["[562]",    "Cuota IS — tipo general 25%",
                         fmt_eur(_cuota_pdf)],
                        ["[582]",    "Retenciones soportadas",
                         f"−{fmt_eur(_retenc_pdf)}"],
                        ["[599]",    "CUOTA DIFERENCIAL (a ingresar)",
                         fmt_eur(_dif_pdf)],
                    ]
                    _hdr_color = colors.HexColor("#059669")
                    _bas_color = colors.HexColor("#E6F9F1")
                    _bas_txt   = colors.HexColor("#059669")
                    _bas_row   = -4  # [399]
                else:
                    _el.append(Paragraph("Modelo 100 consolidado", p_h2))
                    _gr = [
                        ["Casilla", "Concepto", "Importe"],
                        ["0102",    "Ingresos totales",
                         fmt_eur(modelo.get("ingresos",0))],
                        ["0105-0112","Gastos deducibles",
                         f"−{fmt_eur(modelo.get('total_gastos',0))}"],
                        ["0149",    "Rendimiento neto",
                         fmt_eur(modelo.get("rend_neto",0))],
                        ["0150",    "Reduccion aplicada",
                         f"−{fmt_eur(modelo.get('reduccion',0))}"],
                        ["0156",    "BASE IMPONIBLE ESTIMADA",
                         fmt_eur(modelo.get("rend_final",0))],
                    ]
                    _hdr_color = _PU
                    _bas_color = _LP
                    _bas_txt   = _PU
                    _bas_row   = -1

                _gt = Table(_gr, colWidths=[3*cm, 9.5*cm, 4*cm])
                _gt.setStyle(TableStyle([
                    ("BACKGROUND",     (0,0),(-1,0),  _hdr_color),
                    ("TEXTCOLOR",      (0,0),(-1,0),  colors.white),
                    ("FONTNAME",       (0,0),(-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",       (0,0),(-1,-1), 8),
                    ("ROWBACKGROUNDS", (0,1),(-1,-2), [_LG, colors.white]),
                    ("BACKGROUND",     (0,_bas_row),(-1,_bas_row), _bas_color),
                    ("FONTNAME",       (0,_bas_row),(-1,_bas_row), "Helvetica-Bold"),
                    ("TEXTCOLOR",      (0,_bas_row),(-1,_bas_row), _bas_txt),
                    ("GRID",           (0,0),(-1,-1), 0.3,colors.HexColor("#E2E8F0")),
                    ("ALIGN",          (2,0),(2,-1),  "RIGHT"),
                    ("PADDING",        (0,0),(-1,-1), 5),
                ]))
                _el.append(_gt)
                _el.append(Spacer(1, 10))

                # ── Desglose por inmueble ─────────────────────────
                _el.append(Paragraph("Desglose por inmueble", p_h2))
                if es_soc_gl:
                    _ir = [["Inmueble", "[318] Ingresos", "[319] Gastos", "[399] Resultado", "[562] IS 25%"]]
                else:
                    _ir = [["Inmueble", "Ingresos", "Gastos", "Rend. neto", "Base imp."]]
                for _, _row in df_inm.iterrows():
                    _nm2 = str(_row.get(col_n,""))
                    _mi2 = calcular_modelo100_inmueble(_row, df_mov)
                    _sem2 = calcular_semaforo_inmueble(_row)
                    _n_al = len(_sem2.get("problemas",[]))
                    _nm2_disp = _nm2[:26] + (f" ({_n_al} alertas)" if _n_al else "")
                    if es_soc_gl:
                        _ing2 = _mi2.get("0102", _mi2.get("ingresos", 0))
                        _gas2 = _mi2.get("0107", _mi2.get("total_gastos", 0))
                        _res2 = _ing2 - _gas2
                        _is2  = round(max(_res2 * 0.25, 0), 2)
                        _ir.append([_nm2_disp, fmt_eur(_ing2), fmt_eur(_gas2),
                                    fmt_eur(_res2), fmt_eur(_is2)])
                    else:
                        _ir.append([
                            _nm2_disp,
                            fmt_eur(_mi2.get("ingresos",0)),
                            fmt_eur(_mi2.get("total_gastos",0)),
                            fmt_eur(_mi2.get("rend_neto",0)),
                            fmt_eur(_mi2.get("rend_final",0)),
                        ])
                _it = Table(_ir, colWidths=[5.5*cm,3*cm,2.5*cm,3*cm,2.5*cm])
                _it.setStyle(TableStyle([
                    ("BACKGROUND",     (0,0),(-1,0),  colors.HexColor("#059669") if es_soc_gl else _LG),
                    ("TEXTCOLOR",      (0,0),(-1,0),  colors.white if es_soc_gl else _DK),
                    ("FONTNAME",       (0,0),(-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",       (0,0),(-1,-1), 8),
                    ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.HexColor("#F8FAFC"), colors.white]),
                    ("GRID",           (0,0),(-1,-1), 0.3,colors.HexColor("#E2E8F0")),
                    ("ALIGN",          (1,0),(-1,-1), "RIGHT"),
                    ("PADDING",        (0,0),(-1,-1), 5),
                ]))
                _el.append(_it)

                _el.append(Spacer(1, 16))
                _el.append(HRFlowable(width="100%", thickness=0.5,
                    color=_GR, spaceAfter=4))
                _nota_gl = (
                    "Documento orientativo generado por FiscalHub. "
                    "Sociedad Patrimonial | IS tipo general 25% (Art. 29 LIS). "
                    "Sin reduccion por arrendamiento de vivienda habitual. "
                    "Verificar con software oficial AEAT antes de presentar el Modelo 200."
                ) if es_soc_gl else (
                    "Documento orientativo generado por FiscalHub. "
                    "Base imponible estimada al tipo marginal del 30%. "
                    "Verificar con software oficial AEAT antes de presentar."
                )
                _el.append(Paragraph(_nota_gl, p_sm))

                _doc.build(_el)
                _buf.seek(0)
                st.session_state["fh_gl_pdf"] = _buf.read()
            except Exception as _e:
                st.error(f"Error generando PDF: {str(_e)[:200]}")

        if "fh_gl_pdf" in st.session_state:
            st.download_button("⬇️ Descargar PDF",data=st.session_state["fh_gl_pdf"],
                file_name=f"{'IS_Modelo200' if es_soc_gl else 'IRPF'}_{nombre.replace(' ','_')}_2025_global.pdf",
                mime="application/pdf",use_container_width=True,key="gl_pdf_dl")
    with e2:
        if st.button("📊 Exportar Excel", use_container_width=True, key="gl_xlsx"):
            try:
                import openpyxl
                from openpyxl.styles import (Font, PatternFill, Alignment,
                                              Border, Side)
                from openpyxl.utils import get_column_letter
                import io as _io2

                wb  = openpyxl.Workbook()
                ws  = wb.active
                ws.title = "Modelo 100"

                _PU_xl = "FF534AB7"
                _GN_xl = "FF059669"
                _RD_xl = "FFDC2626"
                _LG_xl = "FFF0EEFF"
                _GR_xl = "FFF1F5F9"

                def _hdr_cell(cell, txt, bg=_PU_xl):
                    cell.value = txt
                    cell.font  = Font(bold=True, color="FFFFFFFF", size=10)
                    cell.fill  = PatternFill("solid", fgColor=bg)
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center", wrap_text=True)

                def _data_cell(cell, val, bold=False, color="FF0F172A", align="left"):
                    cell.value = val
                    cell.font  = Font(bold=bold, color=color, size=9)
                    cell.alignment = Alignment(horizontal=align, vertical="center")

                thin = Side(style="thin", color="FFE2E8F0")
                brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

                # ── Cabecera ──────────────────────────────────────
                ws.merge_cells("A1:G1")
                c = ws["A1"]
                c.value = f"FiscalHub · Resumen IRPF 2025 — {nombre}"
                c.font  = Font(bold=True, size=13, color="FF534AB7")
                c.alignment = Alignment(horizontal="center")
                ws.row_dimensions[1].height = 22

                ws.merge_cells("A2:G2")
                ws["A2"].value = (f"Asesor: {nombre_asesor or '—'} · "
                                  f"Generado: {_d2.today().strftime('%d/%m/%Y')}")
                ws["A2"].font  = Font(size=8, color="FF64748B")
                ws["A2"].alignment = Alignment(horizontal="center")
                ws.row_dimensions[2].height = 14

                # ── Cabecera columnas ─────────────────────────────
                hdrs = ["Inmueble","Tipo","Inquilino",
                        "Ingresos €","Gastos €","Rend. neto €",
                        "Base imp. €"]
                for ci, h in enumerate(hdrs, 1):
                    _hdr_cell(ws.cell(4, ci), h)
                ws.row_dimensions[4].height = 18

                # ── Filas por inmueble ────────────────────────────
                totales_i = totales_g = totales_n = totales_b = 0
                for ri, (_, row_x) in enumerate(df_inm.iterrows(), 5):
                    nm_x  = str(row_x.get(col_n,""))
                    m_x   = calcular_modelo100_inmueble(row_x, df_mov)
                    tp_x  = str(row_x.get("tipo_arrendamiento") or "—")
                    iq_x  = str(row_x.get("inquilino") or "—")
                    bg_row = _GR_xl if ri % 2 == 0 else "FFFFFFFF"
                    for ci_ in range(1, 8):
                        ws.cell(ri, ci_).fill = PatternFill("solid", fgColor=bg_row)
                        ws.cell(ri, ci_).border = brd

                    _data_cell(ws.cell(ri,1), nm_x, bold=True)
                    _data_cell(ws.cell(ri,2), tp_x)
                    _data_cell(ws.cell(ri,3), iq_x)
                    _data_cell(ws.cell(ri,4), m_x["ingresos"],
                               color=_GN_xl, align="right")
                    _data_cell(ws.cell(ri,5), -m_x["total_gastos"],
                               color=_RD_xl, align="right")
                    _data_cell(ws.cell(ri,6), m_x["rend_neto"], align="right")
                    _data_cell(ws.cell(ri,7), m_x["rend_final"],
                               bold=True, color="FF534AB7", align="right")
                    totales_i += m_x["ingresos"]
                    totales_g += m_x["total_gastos"]
                    totales_n += m_x["rend_neto"]
                    totales_b += m_x["rend_final"]
                    ws.row_dimensions[ri].height = 16

                # ── Fila totales ──────────────────────────────────
                tr = len(list(df_inm.iterrows())) + 5
                _hdr_cell(ws.cell(tr,1), "TOTAL", bg=_PU_xl)
                for ci_ in range(2,4):
                    ws.cell(tr,ci_).fill = PatternFill("solid", fgColor=_PU_xl)
                _data_cell(ws.cell(tr,4), totales_i,
                           bold=True, color="FFFFFFFF", align="right")
                ws.cell(tr,4).fill = PatternFill("solid", fgColor=_PU_xl)
                _data_cell(ws.cell(tr,5), -totales_g,
                           bold=True, color="FFFFFFFF", align="right")
                ws.cell(tr,5).fill = PatternFill("solid", fgColor=_PU_xl)
                _data_cell(ws.cell(tr,6), totales_n,
                           bold=True, color="FFFFFFFF", align="right")
                ws.cell(tr,6).fill = PatternFill("solid", fgColor=_PU_xl)
                _data_cell(ws.cell(tr,7), totales_b,
                           bold=True, color="FFFFFFFF", align="right")
                ws.cell(tr,7).fill = PatternFill("solid", fgColor=_PU_xl)
                ws.row_dimensions[tr].height = 18

                # Ancho columnas
                for ci_, w_ in enumerate([28,18,18,14,14,14,14], 1):
                    ws.column_dimensions[get_column_letter(ci_)].width = w_

                ws.freeze_panes = "A5"

                _xbuf = _io2.BytesIO()
                wb.save(_xbuf)
                _xbuf.seek(0)
                st.session_state["fh_gl_xlsx"] = _xbuf.read()
            except Exception as _ex:
                st.error(f"Error Excel: {str(_ex)[:150]}")
        if "fh_gl_xlsx" in st.session_state:
            st.download_button("⬇️ Descargar Excel",
                data=st.session_state["fh_gl_xlsx"],
                file_name=f"IRPF_{nombre.replace(' ','_')}_2025.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="gl_xlsx_dl")


# ── ALERTAS ───────────────────────────────────────────────────────
def pantalla_alertas():
    cartera = st.session_state.get("fh_cartera", [])
    todas = []
    for c in cartera:
        for a in c.get("alertas",[]):
            todas.append({**a, "cliente_nombre": c["nombre"], "cliente_id": c["id"]})
    todas.sort(key=lambda x:(0 if x["tipo"]=="crit" else 1))
    n_cr = len([a for a in todas if a["tipo"]=="crit"])
    n_wn = len([a for a in todas if a["tipo"]=="warn"])

    st.markdown(f"""<div style="margin-bottom:16px;">
      <div class="nc-page-label">Cartera completa · por urgencia</div>
      <div class="nc-page-title">Alertas fiscales</div>
      <div class="nc-page-sub">{len(todas)} alertas · {n_cr} críticas · {n_wn} a revisar</div>
    </div>""", unsafe_allow_html=True)

    imp = sum(a.get("impacto",0) for a in todas if a.get("impacto",0)>0)
    render_kpi_row([
        {"label":"🚨 Críticas",   "value":str(n_cr),
         "color":RED,    "subtitle":"Antes del 30 jun"},
        {"label":"⚡ A revisar",  "value":str(n_wn),
         "color":AMBER,  "subtitle":"Esta semana"},
        {"label":"💶 Impacto",    "value":fmt_eur(imp),
         "color":ACCENT_F, "subtitle":"Recuperable"},
    ])

    st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)
    if not todas: st.success("✅ Sin alertas activas."); return

    # Separar críticas y medias
    criticas = [a for a in todas if a["tipo"] == "crit"]
    medias   = [a for a in todas if a["tipo"] == "warn"]

    def _render_alertas_grid(lista, seccion_key):
        """Cards de alerta con patrón header coloreado + body + botón clicable."""
        MAX_COLS = 4
        for fila_start in range(0, len(lista), MAX_COLS):
            fila_rows = lista[fila_start:fila_start+MAX_COLS]
            cols = st.columns(MAX_COLS)
            for col_idx, a in enumerate(fila_rows):
                es_crit = a["tipo"] == "crit"
                tipo_lbl = "⚠️ Crítica" if es_crit else "◔ Revisar"
                # Color header: color del cliente (determinista)
                cli_id  = a.get("cliente_id", a.get("cliente_nombre",""))
                hdr_col = _color_cli(cli_id)
                # Badge color por tipo
                badge_bg = "rgba(220,38,38,0.12)" if es_crit else "rgba(217,119,6,0.12)"
                badge_col= "#DC2626" if es_crit else "#D97706"
                nm      = _e(a.get("cliente_nombre",""))
                inm     = _e(a.get("inmueble","")[:40])
                titulo  = _e(a.get("titulo",""))
                desc    = _e(a.get("desc","")[:120])
                accion  = _e(a.get("accion","")[:80])

                with cols[col_idx]:
                    html = (
                        f'<div style="background:{hdr_col};border-radius:12px 12px 0 0;'
                        f'padding:12px 14px 10px;margin-bottom:-1px;">'
                        f'<div style="font-size:11px;color:rgba(255,255,255,0.65);'
                        f'font-weight:600;letter-spacing:0.05em;margin-bottom:3px;">'
                        f'{tipo_lbl}</div>'
                        f'<div style="font-size:15px;font-weight:800;color:#FFF;'
                        f'line-height:1.2;">{nm}</div>'
                        f'<div style="font-size:11px;color:rgba(255,255,255,0.65);'
                        f'margin-top:2px;">📍 {inm}</div>'
                        f'</div>'
                        f'<div style="background:#FFF;border:2px solid #E2E8F0;'
                        f'border-top:none;border-radius:0 0 12px 12px;'
                        f'padding:12px 14px 10px;">'
                        f'<div style="font-size:14px;font-weight:700;color:#1e293b;'
                        f'margin-bottom:4px;">{titulo}</div>'
                        f'<div style="font-size:12px;color:#64748B;margin-bottom:8px;'
                        f'line-height:1.4;">{desc}</div>'
                        f'<div style="background:{badge_bg};border-radius:6px;'
                        f'padding:5px 10px;font-size:11px;font-weight:700;'
                        f'color:{badge_col};">→ {accion}</div>'
                        f'</div>'
                    )
                    st.markdown(html, unsafe_allow_html=True)
                    # Botón clicable — navega al cliente
                    cli_obj = next((c for c in cartera
                                    if c["nombre"]==a.get("cliente_nombre","")), None)
                    if st.button("🔍 Ir al cliente",
                                 key=f"alt_{seccion_key}_{fila_start}_{col_idx}",
                                 use_container_width=True):
                        if cli_obj:
                            st.session_state.fh_cliente_sel = cli_obj["id"]
                            st.session_state.fh_menu = "cliente"
                            st.rerun()
                    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    if criticas:
        st.markdown('<div class="nc-section">🔴 Críticas — acción urgente</div>',
                    unsafe_allow_html=True)
        _render_alertas_grid(criticas, "cr")

    if medias:
        st.markdown('<div class="nc-section" style="margin-top:20px;">🟡 A revisar esta semana</div>',
                    unsafe_allow_html=True)
        _render_alertas_grid(medias, "wn")

# ── EXPORTAR ──────────────────────────────────────────────────────
def pantalla_exportar():
    cartera = st.session_state.get("fh_cartera",[])
    st.markdown("""<div style="margin-bottom:20px;">
      <div class="nc-page-label">Generación de entregables</div>
      <div class="nc-page-title">Exportar documentos</div>
      <div class="nc-page-sub">Revisa y exporta el modelo 100 de cada cliente.</div>
    </div>""", unsafe_allow_html=True)

    if not cartera:
        st.info("Sin clientes vinculados.")
        return

    _HDR = {"critico":"#7F1D1D","medio":"#78350F","ok":"#14532D"}
    _COL = {"critico":"#DC2626","medio":"#D97706","ok":"#059669"}
    _LBL = {"critico":"⚠ Crítico","medio":"◔ Revisar","ok":"✓ OK"}

    MAX_COLS = 4
    for fila_start in range(0, len(cartera), MAX_COLS):
        fila_rows = cartera[fila_start:fila_start+MAX_COLS]
        cols = st.columns(MAX_COLS)
        for col_idx, c in enumerate(fila_rows):
            estado  = c["estado"]
            hdr     = _HDR[estado]
            txt     = _COL[estado]
            lbl     = _LBL[estado]
            modelo  = c.get("modelo100",{})
            ingresos= fmt_eur(modelo.get("ingresos",0))
            base    = fmt_eur(modelo.get("rend_final",0))
            gastos  = fmt_eur(modelo.get("total_gastos",0))
            rend    = fmt_eur(modelo.get("rend_neto",0))
            criticas= c["criticas"]
            badge_bg= {"critico":"rgba(220,38,38,0.10)",
                       "medio":"rgba(217,119,6,0.10)",
                       "ok":"rgba(5,150,105,0.10)"}[estado]
            chk = "✅" if criticas == 0 else "⚠️"

            with cols[col_idx]:
                _es_soc = c.get("tipo_cuenta") == "sociedad"
                _mis    = c.get("modelo_is", {})
                if _es_soc:
                    _label_ing  = "📥 [318] Ingresos"
                    _label_gas  = "📤 [319] Gastos"
                    _label_rend = "⚖️ [399] Base imp."
                    _label_base = "🏛️ [562] Cuota IS 25%"
                    ingresos = fmt_eur(_mis.get("ingresos", 0))
                    gastos   = fmt_eur(_mis.get("total_gastos", 0))
                    rend     = fmt_eur(_mis.get("resultado", 0))
                    base     = fmt_eur(_mis.get("cuota_is", 0))
                    _subtitulo = "IS · Modelo 200"
                else:
                    _label_ing  = "📥 0102 Ingresos"
                    _label_gas  = "📤 Gastos deducibles"
                    _label_rend = "⚖️ 0149 Rend. neto"
                    _label_base = "🧾 Base imp. est."
                    _subtitulo  = "IRPF · Modelo 100"

                _badge_soc = (
                    '<span style="background:rgba(5,150,105,0.15);color:#059669;' +
                    'font-size:10px;font-weight:700;padding:2px 7px;border-radius:4px;' +
                    'margin-left:6px;">🏢 IS</span>'
                ) if _es_soc else ""

                st.markdown(
                    f'<div style="background:{hdr};border-radius:12px 12px 0 0;' +
                    f'padding:14px 16px 12px;display:flex;align-items:center;' +
                    f'gap:10px;margin-bottom:-1px;">' +
                    f'<span style="font-size:22px;">📋</span>' +
                    f'<div style="flex:1;min-width:0;">' +
                    f'<div style="font-size:16px;font-weight:800;color:#FFF;' +
                    f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">' +
                    f'{c["nombre"]}{_badge_soc}</div>' +
                    f'<div style="font-size:12px;color:rgba(255,255,255,0.65);margin-top:2px;">' +
                    f'{c["inmuebles"]} inmuebles · {_subtitulo}</div></div>' +
                    f'<span style="background:rgba(255,255,255,0.15);color:#FFF;' +
                    f'font-size:11px;font-weight:700;padding:3px 8px;' +
                    f'border-radius:6px;">{lbl}</span></div>' +
                    f'<div style="background:#FFF;border:2px solid #E2E8F0;' +
                    f'border-top:none;border-radius:0 0 12px 12px;' +
                    f'padding:14px 16px 12px;">' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:6px;">' +
                    f'<span style="font-size:13px;color:#94A3B8;">{_label_ing}</span>' +
                    f'<span style="font-size:14px;font-weight:800;color:#059669;">{ingresos}</span></div>' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:6px;">' +
                    f'<span style="font-size:13px;color:#94A3B8;">{_label_gas}</span>' +
                    f'<span style="font-size:14px;font-weight:800;color:#DC2626;">-{gastos}</span></div>' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:6px;">' +
                    f'<span style="font-size:13px;color:#94A3B8;">{_label_rend}</span>' +
                    f'<span style="font-size:14px;font-weight:800;color:#534AB7;">{rend}</span></div>' +
                    f'<div style="display:flex;justify-content:space-between;margin-bottom:10px;">' +
                    f'<span style="font-size:13px;color:#94A3B8;">{chk} Alertas</span>' +
                    f'<span style="font-size:14px;font-weight:800;color:{txt};">{criticas}</span></div>' +
                    f'<div style="background:{badge_bg};border-radius:6px;' +
                    f'padding:6px 10px;text-align:center;' +
                    f'font-size:13px;font-weight:700;color:{txt};">' +
                    f'{_label_base}: {base}</div></div>',
                    unsafe_allow_html=True)

                if st.button("📄 Ir a revisión completa",
                             key=f"exp_{c['id']}_{col_idx}",
                             use_container_width=True, type="primary"):
                    st.session_state.fh_cliente_sel = c["id"]
                    st.session_state.fh_menu = "cliente"
                    st.rerun()
                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ── VINCULAR ──────────────────────────────────────────────────────
def pantalla_vincular():
    st.markdown("""<div style="margin-bottom:16px;">
      <div class="nc-page-label">Conectar con Nolasco Capital</div>
      <div class="nc-page-title">Vincular cliente</div>
      <div class="nc-page-sub">Introduce el código de 6 dígitos del propietario.</div>
    </div>""", unsafe_allow_html=True)
    st.markdown("""<div class="nc-callout info" style="max-width:560px;margin-bottom:20px;">
      <strong>¿Cómo funciona?</strong> El propietario entra en Nolasco Capital →
      Privacidad → "Compartir con Asesor" → genera un código. Te lo envía y lo introduces aquí.
    </div>""", unsafe_allow_html=True)
    codigo = st.text_input("Código:", max_chars=10, placeholder="Ej: 628410", key="vincular_codigo")
    if st.button("🔗 Vincular", use_container_width=True, type="primary", key="vincular_btn"):
        if codigo.strip():
            with st.spinner("Verificando..."):
                res = vincular_propietario(st.session_state.fh_user_id, codigo.strip().upper())
            if res["ok"]:
                st.success(f"✅ Vinculado — {res.get('nombre','Propietario')} en tu cartera.")
                vinculos = get_clientes_vinculados(st.session_state.fh_user_id)
                st.session_state.fh_cartera = construir_cartera(vinculos)
                st.rerun()
            else: st.error(f"❌ {res.get('error','Código no válido')}")
        else: st.warning("Introduce un código")


# ── MOCK DATA — activar con ?demo=1 en la URL ─────────────────────
def _mock_cartera():
    """Cartera de demostración con 6 clientes ficticios + los reales."""
    import pandas as pd
    from datetime import date, timedelta

    def _inm(nombre, inquilino, tipo, renta, ibi, amort, seguro, comunidad,
             hipoteca=0, tiene_alerta=False, alerta_tipo="crit", alerta_txt=""):
        return {
            "nombre": nombre, "inquilino": inquilino,
            "tipo_arrendamiento": tipo, "renta": renta,
            "ibi_anual": ibi, "amortizacion_fiscal": amort,
            "seguro_anual": seguro, "comunidad": comunidad,
            "intereses_hipoteca": hipoteca,
            "tiene_alerta": tiene_alerta,
            "alerta_tipo": alerta_tipo, "alerta_txt": alerta_txt,
            "fecha_vencimiento_contrato": str(date.today() + timedelta(days=30)) if tiene_alerta else "",
            "precio_compra": 180000, "valor_catastral": 95000,
            "porcentaje_construccion": 0.7,
        }

    clientes_mock = [
        {
            "id": "mock-001", "nombre": "García Martínez, Ana",
            "estado": "critico", "criticas": 3, "medias": 1,
            "inmuebles": 3, "impacto": -2400,
            "alertas": [
                {"tipo":"crit","titulo":"Amortización a 0 — revisar","desc":"Catastral: 95.000 € · Precio compra: 180.000 €","accion":"Calcular 3% s/ MAX(precio compra, catastral) × % construcción","inmueble":"Calle Mayor 12","cliente_nombre":"García Martínez, Ana"},
                {"tipo":"crit","titulo":"Contrato próximo a vencer","desc":"Vence en 28 días sin renovar","accion":"Notificar al inquilino y preparar nuevo contrato","inmueble":"Av. Constitución 4","cliente_nombre":"García Martínez, Ana"},
                {"tipo":"crit","titulo":"ROE negativo","desc":"La hipoteca consume el 110% de los ingresos netos","accion":"Revisar si refinanciar o vender el activo","inmueble":"Plaza Nueva 8","cliente_nombre":"García Martínez, Ana"},
                {"tipo":"warn","titulo":"Rentabilidad por debajo de mercado","desc":"Rendimiento actual 3.2% vs 6.8% de media en CP 18001","accion":"Evaluar subida de renta en próxima renovación","inmueble":"Calle Mayor 12","cliente_nombre":"García Martínez, Ana"},
            ],
            "modelo100": {"ingresos":28800,"total_gastos":22400,"rend_neto":6400,"reduccion":3200,"rend_final":3200,"retenciones":5472,"intereses":4200,"reparaciones":800,"ibi":1240,"comunidad_seguros":3600,"suministros":0,"gastos_juridicos":0,"amortizacion":12560,"red_pct":50},
            "df_inm": pd.DataFrame([
                _inm("Calle Mayor 12","Luisa Fernández","Larga Duración",900,620,3800,380,1200,1800,True,"crit","Amortización a 0"),
                _inm("Av. Constitución 4","Roberto Sanz","Larga Duración",750,480,2900,290,900,0,True,"crit","Contrato vence 28 días"),
                _inm("Plaza Nueva 8","Carmen López","Larga Duración",550,390,2100,210,720,1600,True,"crit","ROE negativo"),
            ]),
            "df_mov": pd.DataFrame(),
        },
        {
            "id": "mock-002", "nombre": "López Ruiz, Carlos",
            "estado": "critico", "criticas": 2, "medias": 2,
            "inmuebles": 4, "impacto": -1800,
            "alertas": [
                {"tipo":"crit","titulo":"Amortización a 0 — revisar","desc":"Catastral: 112.000 € · Precio compra: 210.000 €","accion":"Calcular 3% s/ MAX","inmueble":"Gran Vía 33","cliente_nombre":"López Ruiz, Carlos"},
                {"tipo":"crit","titulo":"Gastos sin justificar","desc":"3 recibos de reparaciones sin factura adjunta","accion":"Solicitar facturas al propietario antes del 30/06","inmueble":"Recogidas 18","cliente_nombre":"López Ruiz, Carlos"},
                {"tipo":"warn","titulo":"IBI pendiente de actualizar","desc":"El IBI declarado no coincide con el recibo 2024","accion":"Verificar con el Ayuntamiento","inmueble":"Gran Vía 33","cliente_nombre":"López Ruiz, Carlos"},
                {"tipo":"warn","titulo":"Seguro infradeducido","desc":"Seguro hogar+vida deducible al 100% — solo se declaró el 50%","accion":"Corregir casilla 0110","inmueble":"Recogidas 18","cliente_nombre":"López Ruiz, Carlos"},
            ],
            "modelo100": {"ingresos":42000,"total_gastos":31200,"rend_neto":10800,"reduccion":5400,"rend_final":5400,"retenciones":7980,"intereses":6800,"reparaciones":1200,"ibi":1820,"comunidad_seguros":4800,"suministros":0,"gastos_juridicos":0,"amortizacion":16580,"red_pct":50},
            "df_inm": pd.DataFrame([
                _inm("Gran Vía 33","Marcos Vega","Larga Duración",1200,820,5200,480,1560,2400,True,"crit","Amortización a 0"),
                _inm("Recogidas 18","Sofía Moreno","Larga Duración",950,640,3900,360,1200,0,True,"crit","Gastos sin justificar"),
                _inm("Camino Ronda 5","Pedro Jiménez","Temporada",700,420,2800,260,840,0,False),
                _inm("Arabial 22","Nuria Castro","Larga Duración",650,380,2400,220,720,0,False),
            ]),
            "df_mov": pd.DataFrame(),
        },
        {
            "id": "mock-003", "nombre": "Martínez Peña, Isabel",
            "estado": "medio", "criticas": 0, "medias": 2,
            "inmuebles": 2, "impacto": 0,
            "alertas": [
                {"tipo":"warn","titulo":"Contrato vence en 45 días","desc":"Arrendamiento de larga duración próximo a expirar","accion":"Iniciar proceso de renovación o búsqueda de nuevo inquilino","inmueble":"Paseo Colón 7","cliente_nombre":"Martínez Peña, Isabel"},
                {"tipo":"warn","titulo":"Rentabilidad por debajo de mercado","desc":"Renta 4.1% vs 6.5% de media en CP 18002","accion":"Evaluar incremento según IRAV 2026 (2.47%)","inmueble":"San Juan de Dios 14","cliente_nombre":"Martínez Peña, Isabel"},
            ],
            "modelo100": {"ingresos":19200,"total_gastos":12800,"rend_neto":6400,"reduccion":3200,"rend_final":3200,"retenciones":3648,"intereses":2400,"reparaciones":400,"ibi":980,"comunidad_seguros":2400,"suministros":0,"gastos_juridicos":0,"amortizacion":6620,"red_pct":50},
            "df_inm": pd.DataFrame([
                _inm("Paseo Colón 7","Alberto García","Larga Duración",900,680,3200,300,960,0,True,"warn","Contrato vence 45 días"),
                _inm("San Juan de Dios 14","Elena Ruiz","Larga Duración",700,520,2400,240,720,0,True,"warn","Rentabilidad baja"),
            ]),
            "df_mov": pd.DataFrame(),
        },
        {
            "id": "mock-004", "nombre": "Sánchez Torres, Miguel",
            "estado": "ok", "criticas": 0, "medias": 0,
            "inmuebles": 2, "impacto": 0,
            "alertas": [],
            "modelo100": {"ingresos":22800,"total_gastos":14600,"rend_neto":8200,"reduccion":4100,"rend_final":4100,"retenciones":4332,"intereses":0,"reparaciones":600,"ibi":1100,"comunidad_seguros":3200,"suministros":0,"gastos_juridicos":0,"amortizacion":9700,"red_pct":50},
            "df_inm": pd.DataFrame([
                _inm("Alhambra 3","Rosa Blanco","Larga Duración",1100,760,4200,420,1320,0,False),
                _inm("Zaidín Norte 8","Jesús Molina","Larga Duración",800,580,3100,300,960,0,False),
            ]),
            "df_mov": pd.DataFrame(),
        },
        {
            "id": "mock-005", "nombre": "Fernández Gómez, Laura",
            "estado": "ok", "criticas": 0, "medias": 0,
            "inmuebles": 1, "impacto": 0,
            "alertas": [],
            "modelo100": {"ingresos":9600,"total_gastos":6200,"rend_neto":3400,"reduccion":1700,"rend_final":1700,"retenciones":1824,"intereses":0,"reparaciones":200,"ibi":480,"comunidad_seguros":1200,"suministros":0,"gastos_juridicos":0,"amortizacion":4320,"red_pct":50},
            "df_inm": pd.DataFrame([
                _inm("Neptuno 5","Diana Prieto","Larga Duración",800,480,2400,240,720,0,False),
            ]),
            "df_mov": pd.DataFrame(),
        },
        {
            "id": "mock-006", "nombre": "Romero Díaz, Antonio",
            "estado": "critico", "criticas": 1, "medias": 0,
            "inmuebles": 1, "impacto": -600,
            "alertas": [
                {"tipo":"crit","titulo":"Amortización a 0 — revisar","desc":"Catastral: 78.000 € · Precio compra: 145.000 €","accion":"Calcular 3% s/ MAX","inmueble":"Genil 12","cliente_nombre":"Romero Díaz, Antonio"},
            ],
            "modelo100": {"ingresos":9600,"total_gastos":7800,"rend_neto":1800,"reduccion":900,"rend_final":900,"retenciones":1824,"intereses":0,"reparaciones":300,"ibi":560,"comunidad_seguros":1440,"suministros":0,"gastos_juridicos":0,"amortizacion":5500,"red_pct":50},
            "df_inm": pd.DataFrame([
                _inm("Genil 12","Francisco Ruiz","Larga Duración",800,560,2800,280,840,0,True,"crit","Amortización a 0"),
            ]),
            "df_mov": pd.DataFrame(),
        },
    ]
    return sorted(clientes_mock, key=lambda x:({"critico":0,"medio":1,"ok":2}[x["estado"]],-x["criticas"]))


# ── MAIN ──────────────────────────────────────────────────────────
def main():
    inject_global_css("ficahub")
    if "fh_logged" not in st.session_state: st.session_state.fh_logged = False
    if "fh_menu"   not in st.session_state: st.session_state.fh_menu   = "cartera"

    if not st.session_state.fh_logged:
        pantalla_login(); return

    # Demo mode — añade ?demo=1 a la URL para ver mock data
    demo_mode = st.query_params.get("demo", "0") == "1"

    if "fh_cartera" not in st.session_state:
        with st.spinner("Cargando cartera..."):
            if demo_mode:
                # Cargar reales + mock
                vinculos = get_clientes_vinculados(st.session_state.fh_user_id)
                cartera_real = construir_cartera(vinculos)
                cartera_mock = _mock_cartera()
                # Evitar duplicados por nombre
                nombres_reales = {c["nombre"] for c in cartera_real}
                extra = [c for c in cartera_mock if c["nombre"] not in nombres_reales]
                st.session_state.fh_cartera = cartera_real + extra
            else:
                vinculos = get_clientes_vinculados(st.session_state.fh_user_id)
                st.session_state.fh_cartera = construir_cartera(vinculos)

    if demo_mode:
        st.sidebar.markdown(
            '<div style="background:#FFC107;color:#795548;font-size:10px;font-weight:700;'
            'padding:4px 10px;border-radius:6px;margin:4px 10px;text-align:center;'
            'letter-spacing:0.06em;">🎭 MODO DEMO</div>',
            unsafe_allow_html=True
        )

    with st.sidebar:
        render_sidebar()

    menu = st.session_state.get("fh_menu","cartera")
    st.markdown('<div class="nc-page">', unsafe_allow_html=True)
    if   menu == "cartera":        pantalla_cartera()
    elif menu == "cliente":        pantalla_cliente()
    elif menu == "ficha":          pantalla_ficha_inmueble()
    elif menu == "resumen_global": pantalla_resumen_global()
    elif menu == "alertas":        pantalla_alertas()
    elif menu == "exportar":       pantalla_exportar()
    elif menu == "vincular":       pantalla_vincular()
    elif menu == "perfil":         pantalla_perfil()
    st.markdown("</div>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
